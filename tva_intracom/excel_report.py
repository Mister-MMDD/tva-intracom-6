"""Export du recapitulatif, du detail des ventes et des remboursements au format Excel (.xlsx)."""

from __future__ import annotations

import logging
import re
from datetime import date as _date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import ecb_rates
from .i18n import _ as i18n_
from .models import VatResult
from .oss_export import aggregate_oss_results
from .parsers.amazon.detect import parse_date as _parse_amz_date
from .rates import COUNTRY_NAMES, COUNTRY_CURRENCIES
from .report import ReportSummary, build_report

logger = logging.getLogger(__name__)

_COUNTRY_NAMES_XL = COUNTRY_NAMES

_CENT = Decimal("0.01")

def _round(amount: Decimal) -> Decimal:
    return amount.quantize(_CENT, rounding=ROUND_HALF_UP)


def _home_currency(seller_country: str) -> str:
    """Devise locale du pays d'origine du compte (rates.COUNTRY_CURRENCIES)."""
    return COUNTRY_CURRENCIES.get((seller_country or "FR").upper(), "EUR")


def _currency_format(currency_code: str) -> str:
    return f'#,##0.00 "{currency_code}"'


def _to_home_currency(amount: Decimal, currency_code: str, conv_date: _date) -> Decimal:
    """Convertit un montant (calculé en EUR par le moteur fiscal) vers la devise
    locale du pays d'origine, au taux BCE en vigueur à `conv_date` (taux spot au
    moment de la génération du rapport — indicatif : le montant légalement dû
    reste celui calculé en EUR par le moteur, cf. ca3_report.py / oss_xml.py).
    En cas d'indisponibilité du taux BCE, le montant EUR d'origine est renvoyé
    tel quel plutôt que de faire échouer l'export."""
    if not currency_code or currency_code.upper() == "EUR":
        return amount
    try:
        converted, _rate, _info = ecb_rates.convert_to_currency(
            amount, "EUR", currency_code, conv_date,
        )
        return converted
    except Exception:
        return amount

# Noms complets des pays pour l'affichage dans Excel
def _get_country_name(code: str) -> str:
    # On pourrait traduire COUNTRY_NAMES ici via i18n si on voulait
    # mais pour l'instant on garde la logique existante ou on utilise i18n
    # On va privilégier COUNTRY_NAMES qui est déjà complet.
    return COUNTRY_NAMES.get(code.upper(), code)

_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=12, color="1F497D")
_BOLD_FONT = Font(bold=True, size=11)

# Couleurs de chartes graphiques pour les onglets
_BLUE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_ORANGE_HEADER_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
_LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ALERT_FONT = Font(bold=True, color="9C0006")

_EUR_FORMAT = '#,##0.00 "EUR"'
_PCT_FORMAT = '0.##"%"'


_AUTO_WIDTH_SAMPLE_ROWS = 150  # au-delà, échantillon suffisant pour estimer la largeur (_ColumnWidthTracker)


class _ColumnWidthTracker:
    """Calcule la largeur des colonnes au fil de l'écriture des lignes,
    plutôt qu'en relisant le classeur après coup (`ws.iter_rows()`).

    Nécessaire pour les feuilles écrites en mode séquentiel (compatible
    `write_only=True` — voir chantier de réduction mémoire openpyxl) : une
    feuille write_only ne peut pas être relue, donc plus aucune fonction
    d'auto-largeur ne peut s'appuyer sur `iter_rows()` une fois la bascule
    faite. Même échantillonnage que l'ancien `_auto_width` (en-tête +
    `_AUTO_WIDTH_SAMPLE_ROWS` lignes de données) pour un résultat visuel
    identique.
    """

    __slots__ = ("_widths", "_rows_seen")

    def __init__(self) -> None:
        self._widths: dict[int, int] = {}
        self._rows_seen = 0

    def observe_row(self, values: list) -> None:
        self._rows_seen += 1
        if self._rows_seen > _AUTO_WIDTH_SAMPLE_ROWS + 1:  # +1 pour l'en-tête
            return
        if len(values) <= 1:  # Likely a title or note that shouldn't define column width
            return
        for col_idx, value in enumerate(values, 1):
            if value is None:
                continue
            if isinstance(value, (float, Decimal)):
                val_str = f"{value:,.2f} EUR"
            else:
                val_str = str(value)
            length = len(val_str)
            if length > self._widths.get(col_idx, 0):
                self._widths[col_idx] = length

    def apply(self, ws) -> None:
        for col_idx, length in self._widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = max(length + 4, 12)


def _wcell(ws, value, font=None, fill=None, alignment=None, number_format=None):
    """Construit une cellule stylée prête à être ajoutée via `ws.append(...)`.

    `WriteOnlyCell` fonctionne aussi bien en `Workbook()` normal qu'en
    `Workbook(write_only=True)` (vérifié empiriquement) — utiliser ce helper
    dès maintenant permet de convertir les feuilles une par une sans casser
    le mode normal encore actif, puis de basculer tout le classeur en
    write_only une fois toutes les feuilles converties."""
    cell = WriteOnlyCell(ws, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format
    return cell


class _SequentialSheetWriter:
    """Enveloppe transparente autour d'une feuille `write_only` pour contourner
    une limitation d'openpyxl (vérifiée empiriquement, 3.1.5) : les largeurs de
    colonnes (`column_dimensions[...].width`) et les hauteurs de lignes
    (`row_dimensions[...].height`) ne sont prises en compte à la sauvegarde que
    si elles sont définies AVANT le premier `ws.append()` concerné — sinon
    silencieusement ignorées. Or notre code, comme la quasi-totalité du code
    openpyxl "normal", les définit systématiquement juste APRÈS l'append
    correspondant.

    Stratégie : on retarde d'une ligne l'écriture réelle (on ne sait si une
    hauteur va être posée pour la ligne qu'on vient de "append" qu'au moment
    de l'appel suivant), et pour la largeur des colonnes — qui doit être
    fixée avant TOUT append — on bufferise les `_AUTO_WIDTH_SAMPLE_ROWS`
    premières lignes (même échantillon que `_ColumnWidthTracker`), on
    applique les largeurs calculées, puis on relâche ces lignes avant de
    repasser en flux direct (retardé d'une ligne) pour le reste.

    Aucun des ~100 points d'écriture existants (`ws.append(...)` puis
    `ws.row_dimensions[i].height = ...`) n'a besoin d'être modifié : ce
    wrapper se substitue simplement à l'objet `ws` passé aux fonctions
    `_write_*_tab`.
    """

    def __init__(self, ws) -> None:
        object.__setattr__(self, "_ws", ws)
        object.__setattr__(self, "_tracker", _ColumnWidthTracker())
        object.__setattr__(self, "_buffer", [])       # [[row_num, cells, height], ...] avant fixation des largeurs
        object.__setattr__(self, "_pending", None)    # [row_num, cells, height] en attente (mode direct)
        object.__setattr__(self, "_widths_set", False)
        object.__setattr__(self, "_row_counter", 0)

    # -- Passthrough vers la vraie feuille pour tout le reste de l'API --
    def __getattr__(self, name):
        return getattr(self._ws, name)

    def __setattr__(self, name, value):
        if name in ("_ws", "_tracker", "_buffer", "_pending", "_widths_set", "_row_counter"):
            object.__setattr__(self, name, value)
        else:
            setattr(self._ws, name, value)  # ex. ws.title = "..."

    @property
    def row_dimensions(self):
        return _RowDimProxy(self)

    def append(self, row_cells) -> None:
        self._row_counter += 1
        row_num = self._row_counter
        entry = [row_num, row_cells, None]
        if not self._widths_set:
            values = [getattr(c, "value", c) for c in row_cells]
            self._tracker.observe_row(values)
            self._buffer.append(entry)
            if len(self._buffer) > _AUTO_WIDTH_SAMPLE_ROWS + 1:
                # Le vidage se produit ICI, avant que l'appelant n'ait pu fixer
                # la hauteur de cette toute dernière ligne (ws.row_dimensions[i].height
                # est toujours posé juste APRÈS l'append correspondant) : on la
                # garde donc en attente plutôt que de l'émettre tout de suite,
                # sans quoi sa hauteur serait perdue.
                self._flush_sample(keep_last_pending=True)
        else:
            self._flush_pending()
            self._pending = entry

    def _emit(self, row_num: int, row_cells, height) -> None:
        if height is not None:
            self._ws.row_dimensions[row_num].height = height
        self._ws.append(row_cells)

    def _flush_pending(self) -> None:
        if self._pending is not None:
            self._emit(*self._pending)
            self._pending = None

    def _flush_sample(self, keep_last_pending: bool = False) -> None:
        self._tracker.apply(self._ws)
        self._widths_set = True
        last = self._buffer.pop() if (keep_last_pending and self._buffer) else None
        for row_num, row_cells, height in self._buffer:
            self._emit(row_num, row_cells, height)
        self._buffer = []
        if last is not None:
            self._pending = last

    def set_row_height(self, row_num: int, height) -> None:
        if not self._widths_set and row_num <= len(self._buffer) and self._buffer[row_num - 1][0] == row_num:
            self._buffer[row_num - 1][2] = height
            return
        if self._pending is not None and self._pending[0] == row_num:
            self._pending[2] = height
            return
        # Ne devrait pas arriver avec notre pattern d'écriture (hauteur toujours
        # posée juste après l'append de la même ligne) ; on l'ignore par sécurité
        # plutôt que de lever une exception qui casserait un export entier.

    def apply_column_widths(self, widths: dict[int, int]) -> None:
        """Fixe manuellement les largeurs et désactive le tracker automatique."""
        if self._widths_set:
            return

        for col_idx, length in widths.items():
            self._ws.column_dimensions[get_column_letter(col_idx)].width = max(length + 4, 12)

        self._widths_set = True
        # On vide le buffer vers le mode direct (pending), en gardant la dernière
        # ligne pour permettre la fixation de sa hauteur juste après cet appel.
        last = self._buffer.pop() if self._buffer else None
        for row_num, row_cells, height in self._buffer:
            self._emit(row_num, row_cells, height)
        self._buffer = []
        if last is not None:
            self._pending = last

    def finalize(self) -> None:
        """À appeler après la fin d'écriture d'une feuille (voir export_xlsx).
        `keep_last_pending=False` ici : contrairement au vidage déclenché par
        un débordement de l'échantillon (voir append), l'appelant a fini
        d'écrire la feuille, donc toutes les hauteurs — y compris celle de la
        toute dernière ligne — ont déjà été fixées à ce stade."""
        if not self._widths_set:
            self._flush_sample(keep_last_pending=False)
        self._flush_pending()


class _RowDimProxy:
    __slots__ = ("_writer",)
    def __init__(self, writer) -> None:
        self._writer = writer
    def __getitem__(self, row_num: int):
        return _RowDimEntry(self._writer, row_num)


class _RowDimEntry:
    __slots__ = ("_writer", "_row_num")
    def __init__(self, writer, row_num: int) -> None:
        self._writer = writer
        self._row_num = row_num
    @property
    def height(self):
        return None
    @height.setter
    def height(self, value) -> None:
        self._writer.set_row_height(self._row_num, value)


def _oss_period_totals(
    results: list, refund_results: list | None, period: str
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """Totaux OSS agrégés (HT brut, HT remb, TVA brut, TVA remb), recalculés
    au taux BCE de clôture de période (art. 5 bis Règl. UE 2020/194) plutôt
    qu'au taux du jour de vente figé sur summary.oss_ht_by_country /
    summary.oss_by_country.

    Même méthode et même source (`aggregate_oss_results`) que le tableau de
    bord (declarations.py) et que l'onglet OSS détaillé (_write_oss_tab) —
    à utiliser partout où un total OSS est affiché, pour éviter que le
    récapitulatif diverge de ces deux-là sur les ventes en devise étrangère
    (ex. Suède/SEK, Pologne/PLN) : `summary.oss_by_country` seul ne reflète
    que le taux du jour de la vente, pas celui de clôture de la période
    déclarée. Bug constaté en production le 02/08/2026 (écart de 1,29 € sur
    un fichier de test contenant des ventes SEK) — voir post-mortem.
    """
    _z = Decimal("0.00")
    ht_brut = ht_remb = vat_brut = vat_remb = _z
    agg = aggregate_oss_results(list(results) + list(refund_results or []), period=period)
    for _departure, _by_arrival in agg.items():
        for _arrival, _by_rate in _by_arrival.items():
            for _bucket in _by_rate.values():
                ht_brut += _bucket["ht_vente"]
                ht_remb += _bucket["ht_remb"]
                vat_brut += _bucket["tva_vente"]
                vat_remb += _bucket["tva_remb"]
    return ht_brut, ht_remb, vat_brut, vat_remb


def _write_recap(
        ws,
        summary: ReportSummary,
        hash_totals: dict | None = None,
        seller_country: str = "FR",
        display_currency: str | None = None,
        results: list | None = None,
        refund_results: list | None = None,
        period: str = "",
) -> None:
    ws.title = i18n_("xl_tab_recap")

    # Application forcée des largeurs : on le fait au début pour désactiver le 
    # tracker automatique qui se fait piéger par les formules (Excel affiche
    # un montant long "1 234,56 EUR" alors que la valeur stockée est une 
    # formule courte "=B4+C4").
    if hasattr(ws, "apply_column_widths"):
        ws.apply_column_widths({
            1: 50,  # Libellé
            2: 30, 3: 30, 4: 30,  # CA Brut, Remb, Net
            5: 30, 6: 30, 7: 30,  # TVA Brute, Remb, Net
        })

    ws.append([_wcell(ws, i18n_("xl_recap_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25

    # Devise d'affichage choisie par l'utilisateur (display_currency),
    # retombant sur la devise locale du pays d'origine (home_country) si non
    # fournie : les montants calculés en EUR par le moteur fiscal sont
    # convertis pour affichage, au taux BCE du jour de génération du rapport
    # (voir _to_home_currency).
    _currency = display_currency or _home_currency(seller_country)
    _conv_date = _date.today()
    _fmt_home = _currency_format(_currency)

    def _conv(amount: Decimal) -> Decimal:
        return _to_home_currency(amount, _currency, _conv_date)

    if _currency != "EUR":
        ws.append([_wcell(ws, i18n_("xl_recap_currency_note", currency=_currency, date=_conv_date.isoformat()),
                          font=Font(italic=True, size=9, color="7f7f7f"))])
        ws.row_dimensions[2].height = 16
    else:
        ws.append([])

    # Entêtes de la grille de synthèse
    headers = [
        i18n_("xl_recap_col_indicator"),
        i18n_("xl_recap_col_ca_brut"), i18n_("xl_recap_col_ca_remb"), i18n_("xl_recap_col_ca_net"),
        i18n_("xl_recap_col_tva_brute"), i18n_("xl_recap_col_tva_remb"), i18n_("xl_recap_col_tva_nette")
    ]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in headers])
    ws.row_dimensions[3].height = 22

    _z = Decimal("0.00")
    if results is not None:
        # Même méthode que le dashboard / l'onglet OSS détaillé (taux de
        # clôture de période) — voir docstring de _oss_period_totals().
        oss_ht_brut, oss_ht_remb, oss_vat_brut, oss_vat_remb = _oss_period_totals(
            results, refund_results, period
        )
    else:
        # Comportement historique (fallback CLI / appels sans results) :
        # taux du jour de vente, figé sur summary.oss_*_by_country.
        oss_ht_brut = sum(summary.oss_ht_by_country.values(), _z)
        oss_ht_remb = sum(summary.refund_oss_ht_by_country.values(), _z)
        oss_vat_brut = sum(summary.oss_by_country.values(), _z)
        oss_vat_remb = sum(summary.refund_oss_by_country.values(), _z)

    local_ht_brut = sum(summary.local_ht_by_country.values(), _z)
    local_ht_remb = sum(summary.refund_local_ht_by_country.values(), _z)
    local_vat_brut = sum(summary.local_by_country.values(), _z)
    local_vat_remb = sum(summary.refund_local_by_country.values(), _z)

    # Libellé du poste "domestique pays d'origine"
    if (seller_country or "FR").upper() == "FR":
        _home_label = i18n_("xl_indicator_vat_fr")
    else:
        _home_label = i18n_("xl_indicator_vat_home_generic", country=_get_country_name(seller_country))

    # [Libellé, HT Brut, HT Remb, TVA Brut, TVA Remb]
    data_structure = [
        (i18n_("xl_indicator_ca_ht"),          summary.total_ht,          summary.refund_total_ht,   _z,                       _z),
        (_home_label,                      summary.fr_domestic_ht,    summary.refund_fr_domestic_ht, summary.fr_domestic_vat, summary.refund_fr_domestic_vat),
        (i18n_("xl_indicator_vat_oss"),        oss_ht_brut,               oss_ht_remb,               oss_vat_brut,             oss_vat_remb),
        (i18n_("xl_indicator_vat_amazon"),     summary.amazon_ht,         summary.refund_amazon_ht,  summary.amazon_vat,       summary.refund_amazon_vat),
        (i18n_("xl_indicator_vat_local"),      local_ht_brut,             local_ht_remb,             local_vat_brut,           local_vat_remb),
        (i18n_("xl_indicator_vat_import"),     summary.import_ht,         summary.refund_import_ht,  summary.import_vat,       summary.refund_import_vat),
        (i18n_("xl_indicator_b2b_exempt"),     summary.reverse_charge_ht, summary.refund_reverse_charge_ht, _z,                _z),
        (i18n_("xl_indicator_export_exempt"),   summary.export_ht,         summary.refund_export_ht,  _z,                       _z),
    ]

    current_row = 4
    _row_ca3: int | None = None
    _row_oss: int | None = None
    _row_local: int | None = None

    for idx, (label, ht_brut, ht_remb, vat_brut, vat_remb) in enumerate(data_structure):
        _hb_f = float(_conv(ht_brut))
        _hr_f = float(_conv(ht_remb))
        _vb_f = float(_conv(vat_brut))
        _vr_f = float(_conv(vat_remb))
        
        # Lignes HT uniquement : CA global (0), B2B exonéré (6), Export (7)
        is_ht_only_row = idx in (0, 6, 7)

        # Construction des cellules de la ligne
        row_cells = [
            _wcell(ws, label),
            _wcell(ws, _hb_f, number_format=_fmt_home),
            _wcell(ws, _hr_f, number_format=_fmt_home),
            _wcell(ws, f"=B{current_row}+C{current_row}", number_format=_fmt_home, font=_BOLD_FONT, fill=_LIGHT_GRAY_FILL),
        ]

        if is_ht_only_row:
            # Pour ces lignes, on laisse les colonnes TVA vides
            row_cells.extend([_wcell(ws, None), _wcell(ws, None), _wcell(ws, None)])
        else:
            row_cells.extend([
                _wcell(ws, _vb_f, number_format=_fmt_home),
                _wcell(ws, _vr_f, number_format=_fmt_home),
                _wcell(ws, f"=E{current_row}+F{current_row}", number_format=_fmt_home, font=_BOLD_FONT, fill=_LIGHT_GRAY_FILL),
            ])

        ws.append(row_cells)
        ws.row_dimensions[current_row].height = 18

        if idx == 1: _row_ca3   = current_row
        if idx == 2: _row_oss   = current_row
        if idx == 4: _row_local = current_row

        current_row += 1

    # Ligne de Total final "TVA net à payer par vous"
    ws.append([])
    current_row += 1

    _tva_brute_formula = f"=E{_row_ca3}+E{_row_oss}+E{_row_local}"
    _tva_remb_formula  = f"=F{_row_ca3}+F{_row_oss}+F{_row_local}"

    ws.append([
        _wcell(ws, i18n_("xl_recap_total_remit"), font=_BOLD_FONT),
        _wcell(ws, None), _wcell(ws, None), _wcell(ws, None), # Pas de total CA ici
        _wcell(ws, _tva_brute_formula, number_format=_fmt_home, font=_BOLD_FONT),
        _wcell(ws, _tva_remb_formula, number_format=_fmt_home, font=_BOLD_FONT),
        _wcell(ws, f"=E{current_row}+F{current_row}", number_format=_fmt_home, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL),
    ])
    ws.row_dimensions[current_row].height = 20

    # ── Contrôle de cohérence comptable ────────────────────────────────────
    ws.append([])
    ws.append([])
    current_row += 3
    ws.append([_wcell(ws, i18n_("xl_audit_integrity_title"), font=_TITLE_FONT)])
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    ws.append([_wcell(ws, i18n_("xl_audit_integrity_help"))])
    ws.append([])
    current_row += 2

    # Tableau de vérification (tel que demandé par l'utilisateur)
    _headers_audit = [i18n_("xl_audit_col_indicator"), i18n_("xl_audit_col_control")]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in _headers_audit])
    current_row += 1

    # Ligne : Total CA HT (somme des canaux)
    _bucket_net_ht = float(_conv(summary.net_ht_total))
    ws.append([
        _wcell(ws, i18n_("xl_audit_total_ht"), font=_BOLD_FONT),
        _wcell(ws, _bucket_net_ht, number_format=_fmt_home, font=_BOLD_FONT),
    ])
    current_row += 1

    # Ligne : CA HT net déclaré
    _declared_net_ht = float(_conv(summary.total_ht + summary.refund_total_ht))
    ws.append([
        _wcell(ws, i18n_("xl_audit_declared_net_ht")),
        _wcell(ws, _declared_net_ht, number_format=_fmt_home),
    ])
    current_row += 1

    # Ligne : Écart de réconciliation
    ws.append([
        _wcell(ws, i18n_("xl_audit_reconciliation_gap"), font=_BOLD_FONT),
        _wcell(ws, f"=B{current_row - 1}-B{current_row - 2}", number_format=_fmt_home, font=_BOLD_FONT),
    ])
    current_row += 1

    if hash_totals:
        ws.append([])
        current_row += 1
        ws.append([
            _wcell(ws, i18n_("xl_audit_total_rows")),
            _wcell(ws, hash_totals.get("count", 0), font=Font(name="Courier New")),
        ])
        current_row += 1
        ws.append([
            _wcell(ws, i18n_("xl_audit_file_signature")),
            _wcell(ws, hash_totals.get("id_hash", 0), font=Font(name="Courier New", bold=True)),
        ])


def _write_details_tab(ws, tab_title: str, results_list: List, is_refund_tab: bool = False, display_currency: str = "EUR") -> None:
    ws.title = tab_title

    _fmt_curr = _currency_format(display_currency)
    _conv_date = _date.today()

    # Optimisation (voir investigation perf du 2026-08-04, log_test_7.txt) :
    # _to_home_currency()->convert_to_currency()->convert_to_eur()+get_rate()
    # refait un aller-retour complet (avec formatage d'une string `_info`
    # jetée) à CHAQUE appel — or display_currency/_conv_date ne changent
    # jamais pendant cette boucle. Sur 7172 lignes × 3 conversions/ligne
    # (montant HT, TVA calculée, TVA Amazon), ça mesurait ~1.7s. On calcule
    # le taux UNE SEULE fois ici, et _conv() ne fait plus qu'une
    # multiplication + arrondi directs (même résultat, cf. _to_home_currency
    # pour le détail des cas EUR/HRK/taux indisponible qu'on reproduit ici).
    _is_eur = not display_currency or display_currency.upper() == "EUR"
    _rate_to_display = None if _is_eur else ecb_rates.get_rate(display_currency, _conv_date)

    def _conv(amount: Decimal) -> float:
        if _is_eur or _rate_to_display is None:
            return float(amount)
        return float(_round(amount * _rate_to_display))

    headers = [
        i18n_("xl_col_tx_id"), i18n_("xl_col_date"), i18n_("xl_col_from"), i18n_("xl_col_to"), i18n_("xl_col_buyer_type"),
        i18n_("xl_col_amount_ht"), i18n_("xl_col_scenario"), i18n_("xl_col_vat_country"), i18n_("xl_col_vat_rate"), i18n_("xl_col_vat_amount"),
        i18n_("xl_col_vat_amazon"), i18n_("xl_col_vat_gap"),
        i18n_("xl_col_collector"), i18n_("xl_col_channel"), i18n_("xl_col_note")
    ]

    header_fill = _ORANGE_HEADER_FILL if is_refund_tab else _BLUE_HEADER_FILL
    _header_align = Alignment(horizontal="center", vertical="center")
    _width_tracker = _ColumnWidthTracker()

    _header_cells = [
        _wcell(ws, text, font=_HEADER_FONT_WHITE, fill=header_fill, alignment=_header_align)
        for text in headers
    ]
    ws.append(_header_cells)
    ws.row_dimensions[1].height = 22
    _width_tracker.observe_row(headers)

    _alert_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for i, r in enumerate(results_list, 2):
        # -- SÉCURITÉ : On détecte si on a un objet VatResult complet ou juste un objet Sale --
        if hasattr(r, "sale"):
            # Cas normal : c'est un VatResult
            sale = r.sale
            scenario_val = str(r.scenario.value)
            vat_rate = r.vat_rate
            vat_amount = r.vat_amount
            collector = r.collector.value
            channel = r.channel.value
            note = r.note
        else:
            # Cas dégradé : c'est juste un objet Sale
            sale = r
            scenario_val = "REFUND"
            vat_rate = 0.0
            vat_amount = 0.0
            collector = "N/A"
            channel = "N/A"
            note = "Remboursement (source brute)"

        _amount_ht = _conv(sale.amount_ht)
        # Pays de taxe : disponible sur VatResult, "-" uniquement en mode degrade (Sale brut)
        _vat_country = getattr(r, "vat_country", "-") if hasattr(r, "vat_country") else "-"
        _vat_rate_f = float(vat_rate)
        _vat_amount_f = _conv(vat_amount)
        _amz_vat = _conv(getattr(sale, "amazon_vat_amount", Decimal("0")))
        _ecart = round(_amz_vat - _vat_amount_f, 2)

        _row_values = [
            str(getattr(sale, "display_id", "") or sale.sale_id),
            str(sale.transaction_date),
            str(sale.stock_country),
            str(sale.buyer_country),
            str(sale.buyer_type.value),
            _amount_ht,
            scenario_val,
            _vat_country or "-",
            _vat_rate_f,
            _vat_amount_f,
            _amz_vat,
            _ecart,
            str(collector),
            str(channel),
            str(note),
            ]

        _row_cells = [
            _wcell(ws, _row_values[0]),
            _wcell(ws, _row_values[1]),
            _wcell(ws, _row_values[2]),
            _wcell(ws, _row_values[3]),
            _wcell(ws, _row_values[4]),
            _wcell(ws, _row_values[5], number_format=_fmt_curr),
            _wcell(ws, _row_values[6]),
            _wcell(ws, _row_values[7]),
            _wcell(ws, _row_values[8], number_format=_PCT_FORMAT),
            _wcell(ws, _row_values[9], number_format=_fmt_curr),
            _wcell(ws, _row_values[10], number_format=_fmt_curr),
            # Colorier en rouge si ecart significatif (> 0.05 EUR)
            _wcell(ws, _row_values[11], number_format=_fmt_curr,
                   fill=_alert_fill if abs(_ecart) > 0.05 else None),
            _wcell(ws, _row_values[12]),
            _wcell(ws, _row_values[13]),
            _wcell(ws, _row_values[14]),
        ]
        ws.append(_row_cells)
        ws.row_dimensions[i].height = 18
        _width_tracker.observe_row(_row_values)

    _width_tracker.apply(ws)


def _write_audit_tab(ws, results: list, vies_affected_sale_ids: set | None = None, vies_summary=None, display_currency: str = "EUR") -> None:
    """Onglet Audit — deux sections :

    1. Réconciliation agrégée : sous-totaux par (nature, pays destination) avec
       écart absolu et % — identifie les catégories systématiquement décalées.
    2. Détail ligne par ligne : chaque vente avec écart > 0.05 € (ou flux GB).
    """
    from collections import defaultdict

    _fmt_curr = _currency_format(display_currency)
    _conv_date = _date.today()
    def _conv(amount: Decimal) -> float:
        return float(_to_home_currency(amount, display_currency, _conv_date))

    vies_affected_sale_ids = vies_affected_sale_ids or set()
    domestic_rc_sale_ids: set[str] = set()
    if vies_summary and hasattr(vies_summary, "reclassifications"):
        for rc in vies_summary.reclassifications:
            if getattr(rc, "is_domestic_reverse_charge", False):
                domestic_rc_sale_ids.add(rc.sale_id)

    def _nature(r) -> str:
        dep = getattr(r.sale, "stock_country", "")
        arr = getattr(r.sale, "buyer_country", "")
        sid = str(r.sale.sale_id)
        tva_amazon = float(getattr(r.sale, "amazon_vat_amount", Decimal("0")))
        tva_moteur = float(r.vat_amount)
        if dep == "GB" or arr == "GB":
            return i18n_("xl_audit_nature_gb")
        if (str(r.sale.sale_id), str(r.sale.amount_ht)) in vies_affected_sale_ids and tva_amazon == 0:
            return i18n_("xl_audit_nature_vies")
        if sid in domestic_rc_sale_ids or (tva_moteur == 0 and tva_amazon > 0 and dep == arr):
            return i18n_("xl_audit_nature_art194")
        return i18n_("xl_audit_nature_taux")

    # ── Section 1 : Réconciliation agrégée ──────────────────────────────
    ws.title = i18n_("xl_tab_audit")
    _width_tracker = _ColumnWidthTracker()

    ws.append([_wcell(ws, i18n_("xl_audit_agg_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 24
    ws.append([_wcell(ws, i18n_("xl_audit_agg_help"), font=Font(italic=True, size=9, color="595959"))])
    ws.row_dimensions[2].height = 18
    ws.append([])
    ws.row_dimensions[3].height = 8

    agg: dict[tuple[str, str], dict] = defaultdict(lambda: {
        "n": 0, "ht": Decimal("0"), "amz": Decimal("0"), "mot": Decimal("0")
    })
    detail_rows = []

    for r in results:
        dep = getattr(r.sale, "stock_country", "")
        arr = getattr(r.sale, "buyer_country", "")
        tva_amz = Decimal(str(round(float(getattr(r.sale, "amazon_vat_amount", Decimal("0"))), 2)))
        tva_mot = Decimal(str(round(float(r.vat_amount), 2)))
        ecart   = tva_amz - tva_mot
        is_gb   = dep == "GB" or arr == "GB"
        nat     = _nature(r)

        if is_gb or abs(float(ecart)) > 0.05:
            agg[(nat, arr)]["n"]   += 1
            agg[(nat, arr)]["ht"]  += r.sale.amount_ht
            agg[(nat, arr)]["amz"] += tva_amz
            agg[(nat, arr)]["mot"] += tva_mot
            detail_rows.append((r, nat, dep, arr, tva_amz, tva_mot, ecart))

    _headers_1 = [
        i18n_("xl_audit_col_nature"), i18n_("xl_audit_col_dest"),
        i18n_("xl_audit_col_count"), i18n_("xl_audit_col_ca_ht"),
        i18n_("xl_audit_col_vat_amz"), i18n_("xl_audit_col_vat_mot"),
        i18n_("xl_audit_col_gap_abs"), i18n_("xl_audit_col_gap_pct"), i18n_("xl_audit_col_risk"),
    ]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in _headers_1])
    ws.row_dimensions[4].height = 22
    _width_tracker.observe_row(_headers_1)

    row = 5
    for (nat, arr), d in sorted(agg.items()):
        ecart_abs = d["amz"] - d["mot"]
        pct = (ecart_abs / d["mot"] * 100) if d["mot"] != 0 else Decimal("0")
        risque = (i18n_("xl_risk_high") if abs(float(pct)) > 10
                  else i18n_("xl_risk_medium") if abs(float(pct)) > 3
        else i18n_("xl_risk_low"))
        _dest_label = f"{_get_country_name(arr)} ({arr})"
        _ht_f, _amz_f, _mot_f, _ecart_f, _pct_f = (
            float(d["ht"]), float(d["amz"]), float(d["mot"]), float(ecart_abs), float(_round(pct))
        )
        _vals1 = [nat, _dest_label, d["n"], _ht_f, _amz_f, _mot_f, _ecart_f, _pct_f, risque]
        ws.append([
            _wcell(ws, nat), _wcell(ws, _dest_label), _wcell(ws, d["n"]),
            _wcell(ws, _conv(d["ht"]), number_format=_fmt_curr),
            _wcell(ws, _conv(d["amz"]), number_format=_fmt_curr),
            _wcell(ws, _conv(d["mot"]), number_format=_fmt_curr),
            _wcell(ws, _conv(ecart_abs), number_format=_fmt_curr,
                   font=Font(bold=True, color="C00000" if abs(float(ecart_abs)) > 1 else "000000")),
            _wcell(ws, _pct_f, number_format='0.0"%"'),
            _wcell(ws, risque),
        ])
        ws.row_dimensions[row].height = 18
        _width_tracker.observe_row(_vals1)
        row += 1

    if row == 5:
        ws.append([_wcell(ws, i18n_("xl_no_gap_detected"), font=Font(italic=True))])
        row = 6

    # ── Section 2 : Détail ligne par ligne ──────────────────────────────
    row += 2
    ws.append([])
    ws.append([])
    ws.append([_wcell(ws, i18n_("xl_audit_detail_title"), font=Font(bold=True, size=11, color="1F497D"))])
    ws.row_dimensions[row].height = 20
    row += 1
    _headers_2 = [
        i18n_("xl_detail_col_sale_id"), i18n_("xl_detail_col_nature"), i18n_("xl_detail_col_flow"),
        i18n_("xl_detail_col_scenario"), i18n_("xl_detail_col_ht"),
        i18n_("xl_detail_col_vat_amz"), i18n_("xl_detail_col_vat_mot"), i18n_("xl_detail_col_gap"),
    ]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in _headers_2])
    ws.row_dimensions[row].height = 22
    _width_tracker.observe_row(_headers_2)
    row += 1

    for r, nat, dep, arr, tva_amz, tva_mot, ecart in detail_rows:
        _flow = f"{dep}→{arr}"
        _vals2 = [
            str(getattr(r.sale, "display_id", "") or r.sale.sale_id), nat, _flow,
            str(r.scenario.value), float(r.sale.amount_ht), float(tva_amz), float(tva_mot), float(ecart),
        ]
        ws.append([
            _wcell(ws, _vals2[0]), _wcell(ws, nat), _wcell(ws, _flow), _wcell(ws, _vals2[3]),
            _wcell(ws, _conv(r.sale.amount_ht), number_format=_fmt_curr),
            _wcell(ws, _conv(tva_amz), number_format=_fmt_curr),
            _wcell(ws, _conv(tva_mot), number_format=_fmt_curr),
            _wcell(ws, _conv(ecart), number_format=_fmt_curr),
        ])
        ws.row_dimensions[row].height = 18
        _width_tracker.observe_row(_vals2)
        row += 1

    if not detail_rows:
        ws.append([_wcell(ws, i18n_("xl_no_line_gap"), font=Font(italic=True))])

    _width_tracker.apply(ws)


def _write_vies_history_tab(ws, results: list, scope_id: str) -> None:
    """Onglet Historique VIES : piste d'audit de chaque vérification effectuée."""
    from .vies_engine import get_vies_history_bulk, normalize_full_vat

    ws.title = i18n_("xl_tab_vies")
    _width_tracker = _ColumnWidthTracker()
    _headers = [
        i18n_("xl_vies_col_vat"), i18n_("xl_vies_col_checked_at"), i18n_("xl_vies_col_status"),
        i18n_("xl_vies_col_country"), i18n_("xl_vies_col_name"), i18n_("xl_vies_col_error")
    ]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in _headers])
    ws.row_dimensions[1].height = 22
    _width_tracker.observe_row(_headers)

    # IMPORTANT (fix onglet vide/incomplet) : vies_check_history.vat_id stocke
    # le numéro NORMALISE (préfixe pays ajouté si absent, cf. normalize_full_vat
    # — la même fonction canonique utilisée par engine.py avant l'appel VIES).
    # buyer_vat_number, lui, est la valeur BRUTE saisie par l'acheteur (ex:
    # "B71547129" sans "ES" pour un NIF espagnol). Interroger l'historique
    # avec la valeur brute ne matche donc que les numéros où l'acheteur avait
    # déjà tapé le préfixe pays complet — tous les autres étaient absents de
    # cet onglet alors qu'ils étaient bel et bien en cache/historique.
    seen_vats: set[str] = set()
    display_by_full_vat: dict[str, str] = {}
    for r in results:
        vat = getattr(r.sale, "buyer_vat_number", "")
        if not vat:
            continue
        full_vat = normalize_full_vat(vat, getattr(r.sale, "buyer_country", ""))
        if not full_vat:
            continue
        seen_vats.add(full_vat)
        # On garde le numéro tel que saisi pour l'affichage (plus lisible /
        # cohérent avec les autres onglets), la clé de recherche reste full_vat.
        display_by_full_vat.setdefault(full_vat, vat)

    history_by_vat = get_vies_history_bulk(scope_id, sorted(seen_vats))

    row = 2
    for full_vat in sorted(seen_vats):
        history = history_by_vat.get(full_vat, [])
        if not history:
            continue
        _display_vat = display_by_full_vat.get(full_vat, full_vat)
        for entry in history:
            _status = i18n_("xl_vies_status_valid") if entry["valid"] else i18n_("xl_vies_status_invalid")
            _vals = [_display_vat, entry["checked_at"], _status, entry["country_code"], entry["name"], entry["error"]]
            ws.append([_wcell(ws, v) for v in _vals])
            ws.row_dimensions[row].height = 16
            _width_tracker.observe_row(_vals)
            row += 1

    if row == 2:
        ws.append([_wcell(ws, i18n_("xl_vies_no_history"))])
    _width_tracker.apply(ws)


def _write_intrastat_tab(
        ws,
        all_fc_transfers: list,
        results: list,
        seller_country: str = "FR",
        display_currency: str = "EUR",
) -> None:
    """Onglet Intrastat / EMEBI (statistique) — aide au remplissage de la déclaration."""
    from .rates import intrastat_emebi_threshold_for_year

    ws.title = i18n_("xl_tab_intrastat")
    GREEN_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    _width_tracker = _ColumnWidthTracker()

    ws.append([_wcell(ws, i18n_("xl_intrastat_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25

    _fmt_curr = _currency_format(display_currency)
    _conv_date = _date.today()
    def _conv(amount: Decimal) -> float:
        return float(_to_home_currency(amount, display_currency, _conv_date))

    # Année de référence pour le seuil : année en cours au moment de la génération.
    _current_year = _date.today().year
    _seuil_annee_ref, _seuil_confirme = intrastat_emebi_threshold_for_year(_current_year)
    _seuil_warning = (
        "" if _seuil_confirme else
        i18n_("xl_intrastat_unconfirmed_warning", year=_current_year)
    )

    # Note légale. Fusion de cellules (colonnes 1-13) abandonnée : incompatible
    # avec le mode write_only. Le texte déborde naturellement sur les cellules
    # vides adjacentes — rendu visuel quasi identique pour un texte d'une ligne.
    _note_text = i18n_("xl_intrastat_note", seller_country=seller_country, year=_current_year, threshold=_seuil_annee_ref, warning=_seuil_warning)
    ws.append([_wcell(ws, _note_text, font=Font(italic=True, size=10, color="C00000"))])
    ws.row_dimensions[2].height = 30
    ws.append([])
    ws.row_dimensions[3].height = 8

    # Calcul du prix moyen HT par ASIN
    asin_avg = _build_asin_avg_price(results)

    # Agrégation des transferts par (départ, arrivée, ASIN, mois)
    from collections import defaultdict
    flux: dict[tuple, dict] = defaultdict(lambda: {"qty": 0, "nb": 0, "designation": ""})
    for t in all_fc_transfers:
        tx_id_unused, date_str, asin, designation, dep, arr, qty = _parse_fc_transfer(t)
        if not dep or not arr:
            continue
        mois = date_str[:7] if date_str else "—"
        key = (dep, arr, asin, mois)
        flux[key]["qty"]          += qty
        flux[key]["nb"]           += 1
        flux[key]["designation"]  = flux[key]["designation"] or designation

    # ── Jauge de seuil annuel (EMEBI) ───────────────────────────────────
    seuil_par_annee: dict[str, dict] = defaultdict(lambda: {"intro": Decimal("0"), "expe": Decimal("0")})
    for (dep, arr, asin, mois), data in flux.items():
        annee = mois[:4] if mois and mois != "—" else "—"
        avg = asin_avg.get(asin, Decimal("0"))
        valeur = _round(Decimal(str(data["qty"])) * avg) if avg else Decimal("0")
        if arr == seller_country:
            seuil_par_annee[annee]["intro"] += valeur
        if dep == seller_country:
            seuil_par_annee[annee]["expe"] += valeur

    current_row = 4
    if seuil_par_annee:
        ws.append([_wcell(ws, i18n_("xl_intrastat_seuil_title"), font=Font(bold=True, size=11, color="C00000"))])
        current_row += 1
        _headers_seuil = [
            i18n_("xl_intrastat_col_year"), i18n_("xl_intrastat_col_sens"), i18n_("xl_intrastat_col_cumul"),
            i18n_("xl_intrastat_col_threshold"), i18n_("xl_intrastat_col_pct"), i18n_("xl_intrastat_col_status"),
        ]
        ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE,
                          fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
                          alignment=Alignment(horizontal="center", vertical="center"))
                   for t in _headers_seuil])
        _width_tracker.observe_row(_headers_seuil)
        current_row += 1
        any_unconfirmed = False
        for annee in sorted(seuil_par_annee):
            try:
                seuil_annee, confirme = intrastat_emebi_threshold_for_year(int(annee))
            except ValueError:
                seuil_annee, confirme = _seuil_annee_ref, _seuil_confirme
            any_unconfirmed = any_unconfirmed or not confirme
            for sens_label, key_sens in [(i18n_("xl_intrastat_introductions"), "intro"), (i18n_("xl_intrastat_dispatches"), "expe")]:
                cumul = seuil_par_annee[annee][key_sens]
                pct = float(cumul / seuil_annee * 100) if seuil_annee else 0.0
                statut = (i18n_("xl_intrastat_status_exceeded") if pct >= 100
                          else i18n_("xl_intrastat_status_near") if pct >= 80
                else i18n_("xl_intrastat_status_ok"))
                if not confirme:
                    statut += i18n_("xl_intrastat_status_unconfirmed")
                _pct_r = round(pct, 1)
                _vals_seuil = [annee, sens_label, float(cumul), float(seuil_annee), _pct_r, statut]
                ws.append([
                    _wcell(ws, annee), _wcell(ws, sens_label),
                    _wcell(ws, _conv(cumul), number_format=_fmt_curr),
                    _wcell(ws, _conv(seuil_annee), number_format=_fmt_curr),
                    _wcell(ws, _pct_r, number_format='0.0"%"',
                           font=Font(bold=True, color="C00000" if pct >= 100 else ("ED7D31" if pct >= 80 else "375623"))),
                    _wcell(ws, statut),
                ])
                ws.row_dimensions[current_row].height = 18
                _width_tracker.observe_row(_vals_seuil)
                current_row += 1
        ws.append([_wcell(ws, i18n_("xl_intrastat_footer", unconfirmed=(i18n_("xl_intrastat_unconfirmed_footer") if any_unconfirmed else "")),
                          font=Font(italic=True, size=9, color="7f7f7f"))])
        current_row += 2
    else:
        ws.append([_wcell(ws, i18n_("xl_intrastat_no_transfer"), font=Font(italic=True))])
        current_row += 2

    # ── Détail introductions / expéditions (UE → seller_country) ────────
    for flow_label_key, is_intro in [
        ("xl_intrastat_intro_label", True),
        ("xl_intrastat_expe_label", False),
    ]:
        ws.append([_wcell(ws, i18n_(flow_label_key, country=seller_country), font=Font(bold=True, size=11, color="375623"))])
        current_row += 1
        _headers_flux = [
            i18n_("xl_intrastat_col_period"), i18n_("xl_intrastat_col_origin"), i18n_("xl_intrastat_col_dest_cc"),
            i18n_("xl_intrastat_col_flow_code"), i18n_("xl_intrastat_col_nature_tx"),
            i18n_("xl_intrastat_col_asin"), i18n_("xl_intrastat_col_desc"),
            i18n_("xl_intrastat_col_cn8"), i18n_("xl_intrastat_col_qty"), i18n_("xl_intrastat_col_mass"),
            i18n_("xl_intrastat_col_val_stat"), i18n_("xl_intrastat_col_delivery"), i18n_("xl_intrastat_col_remark"),
        ]
        ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=GREEN_FILL,
                          alignment=Alignment(horizontal="center", vertical="center"))
                   for t in _headers_flux])
        ws.row_dimensions[current_row].height = 22
        _width_tracker.observe_row(_headers_flux)
        current_row += 1

        rows_written = 0
        sens = i18n_("Intro") if is_intro else i18n_("Expé")
        for (dep, arr, asin, mois), data in sorted(flux.items()):
            if is_intro and arr != seller_country:
                continue
            if not is_intro and dep != seller_country:
                continue

            qty    = data["qty"]
            desc   = data["designation"][:80] if data["designation"] else ""
            avg    = asin_avg.get(asin, Decimal("0"))
            valeur = _round(Decimal(str(qty)) * avg) if avg else Decimal("0")

            _vals_flux = [
                mois, f"{_get_country_name(dep)} ({dep})", f"{_get_country_name(arr)} ({arr})", sens,
                i18n_("xl_intrastat_transfer_desc"), asin, desc, i18n_("xl_intrastat_to_complete"), qty,
                i18n_("xl_intrastat_to_complete"), float(valeur), "DAP / DDP", i18n_("xl_intrastat_estimated_val_remark"),
            ]
            ws.append([
                _wcell(ws, _vals_flux[0]), _wcell(ws, _vals_flux[1]), _wcell(ws, _vals_flux[2]),
                _wcell(ws, _vals_flux[3]), _wcell(ws, _vals_flux[4]), _wcell(ws, _vals_flux[5]),
                _wcell(ws, _vals_flux[6]), _wcell(ws, _vals_flux[7]), _wcell(ws, _vals_flux[8]),
                _wcell(ws, _vals_flux[9]), _wcell(ws, _vals_flux[10], number_format=_EUR_FORMAT),
                _wcell(ws, _vals_flux[11]), _wcell(ws, _vals_flux[12]),
            ])
            ws.row_dimensions[current_row].height = 18
            _width_tracker.observe_row(_vals_flux)
            current_row += 1
            rows_written += 1

        if rows_written == 0:
            ws.append([_wcell(ws, i18n_("xl_intrastat_no_flow_detected", sens=sens))])
            current_row += 1
        current_row += 2

    _width_tracker.apply(ws)


def _next_working_day(d: _date) -> _date:
    """Retourne d si c'est un jour ouvrable, sinon le lundi suivant."""
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def _deadline_oss(ref_date: _date) -> _date:
    """Délai OSS : fin du mois suivant la fin du trimestre."""
    q_end_month = ((ref_date.month - 1) // 3 * 3) + 3  # dernier mois du trimestre courant
    year = ref_date.year
    if q_end_month > 12:
        q_end_month -= 12
        year += 1
    # Fin du mois suivant
    if q_end_month == 12:
        return _date(year + 1, 1, 31)
    elif q_end_month + 1 == 12:
        return _date(year, 12, 31)
    else:
        # Dernier jour du mois suivant
        import calendar
        last_day = calendar.monthrange(year, q_end_month + 1)[1]
        return _date(year, q_end_month + 1, last_day)


def _write_calendar_tab(
        ws,
        results: list,
        all_fc_transfers: list,
        period: str = "",
        seller_country: str = "FR",
) -> None:
    """Onglet Calendrier Fiscal — prochaines échéances déduites des données traitées."""
    ws.title = i18n_("xl_tab_calendar")
    PURPLE_FILL = PatternFill(start_color="6B3FA0", end_color="6B3FA0", fill_type="solid")
    GREEN_FILL  = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    ORANGE_FILL = _ORANGE_HEADER_FILL
    RED_FILL    = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    today       = _date.today()

    _width_tracker = _ColumnWidthTracker()
    ws.append([_wcell(ws, i18n_("xl_cal_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25
    ws.append([_wcell(ws, i18n_("xl_cal_meta", date=today.isoformat(), country=seller_country, period=(period or i18n_("xl_cal_unspecified"))),
                      font=Font(italic=True, size=10, color="595959"))])
    ws.row_dimensions[2].height = 20
    ws.append([])
    ws.row_dimensions[3].height = 8

    _headers = [
        i18n_("xl_cal_col_channel"), i18n_("xl_cal_col_obligation"), i18n_("xl_cal_col_ref_period"),
        i18n_("xl_cal_col_deadline"), i18n_("xl_cal_col_remaining"), i18n_("xl_cal_col_status"),
        i18n_("xl_cal_col_portal"), i18n_("xl_cal_col_legal"),
    ]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=PURPLE_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in _headers])
    ws.row_dimensions[4].height = 22
    _width_tracker.observe_row(_headers)

    row = 5

    def _write_row(canal, obligation, periode_ref, deadline, portail, base_legale, fill):
        nonlocal row
        jours = (deadline - today).days
        statut = i18n_("xl_cal_status_upcoming") if jours > 7 else (i18n_("xl_cal_status_urgent") if jours >= 0 else i18n_("xl_cal_status_overdue"))
        _vals = [canal, obligation, periode_ref, deadline.isoformat(), jours, statut, portail, base_legale]
        ws.append([
            _wcell(ws, canal, fill=fill, font=Font(bold=True, color="FFFFFF")),
            _wcell(ws, obligation),
            _wcell(ws, periode_ref),
            _wcell(ws, deadline.isoformat()),
            _wcell(ws, jours, font=Font(bold=True,
                                        color="C00000" if jours < 0 else ("ED7D31" if jours <= 7 else "375623"))),
            _wcell(ws, statut),
            _wcell(ws, portail),
            _wcell(ws, base_legale),
        ])
        ws.row_dimensions[row].height = 18
        _width_tracker.observe_row(_vals)
        row += 1

    # ── 1. OSS ────────────────────────────────────────────────────────────
    import calendar as _cal
    import re as _re

    oss_quarters: list[tuple[int, int]] = []   # liste de (année, trimestre) couverts

    if period:
        p = period.strip().upper().replace("T", "Q")
        m = _re.fullmatch(r"(\d{4})-Q([1-4])", p)
        if m:
            oss_quarters = [(int(m.group(1)), int(m.group(2)))]
        else:
            # Multi-trimestres / annuel → générer tous les trimestres
            yr_m = _re.fullmatch(r"(\d{4})", p)
            if yr_m:
                oss_quarters = [(int(yr_m.group(1)), q) for q in range(1, 5)]

    # Compléter depuis les dates de ventes OSS si période non reconnue
    if not oss_quarters:
        from .models import Scenario
        seen_qy: set[tuple[int, int]] = set()
        for r in results:
            if r.scenario not in (Scenario.OSS_B2C,):
                continue
            d = (r.sale.transaction_date or "")[:10]
            try:
                yr, mo = int(d[:4]), int(d[5:7])
                seen_qy.add((yr, (mo - 1) // 3 + 1))
            except (ValueError, IndexError):
                pass
        oss_quarters = sorted(seen_qy)

    for yr, q in oss_quarters:
        q_end_month = q * 3
        last_q_day  = _date(yr, q_end_month, _cal.monthrange(yr, q_end_month)[1])
        deadline    = _deadline_oss(last_q_day)
        _write_row(
            "OSS",
            i18n_("xl_cal_oss_task"),
            f"T{q} {yr}",
            deadline,
            "guichet-entreprises.fr / portail OSS DGFIP",
            "Art. 369 sexdecies & septdecies Dir. 2006/112/CE",
            BLUE_FILL := _BLUE_HEADER_FILL,
        )

    # ── 2. CA3 (TVA locale France) ────────────────────────────────────────
    from .models import Channel
    ca3_months: set[tuple[int, int]] = set()
    for r in results:
        if r.channel != Channel.FR_DOMESTIC:
            continue
        d = (r.sale.transaction_date or "")[:10]
        try:
            ca3_months.add((int(d[:4]), int(d[5:7])))
        except (ValueError, IndexError):
            pass
    for yr, mo in sorted(ca3_months):
        next_mo = mo + 1 if mo < 12 else 1
        next_yr = yr if mo < 12 else yr + 1
        deadline = _date(next_yr, next_mo, 24)
        if (seller_country or "FR").upper() == "FR":
            _canal_label = "CA3 / TVA FR"
            _task_label = i18n_("xl_cal_ca3_task")
        else:
            _canal_label = f"TVA {seller_country}"
            _task_label = i18n_("xl_cal_home_task", country=seller_country)
        _write_row(
            _canal_label,
            _task_label,
            f"{yr}-{mo:02d}",
            deadline,
            "impots.gouv.fr (espace professionnel) → Déclarer → TVA" if (seller_country or "FR").upper() == "FR"
            else i18n_("xl_cal_local_portal_generic"),
            "Art. 287 CGI — régime normal mensuel" if (seller_country or "FR").upper() == "FR"
            else i18n_("xl_cal_local_legal_generic"),
            ORANGE_FILL,
        )

    # ── 3. Intrastat ─────────────────────────────────────────────────────
    intrastat_months: set[tuple[int, int]] = set()
    for t in all_fc_transfers:
        tx_id_unused, date_str, asin_unused, desc_unused, dep, arr, qty_unused = _parse_fc_transfer(t)
        if not dep or not arr:
            continue
        if dep != seller_country and arr != seller_country:
            continue
        d = (date_str or "")[:10]
        try:
            intrastat_months.add((int(d[:4]), int(d[5:7])))
        except (ValueError, IndexError):
            pass
    for yr, mo in sorted(intrastat_months):
        next_mo = mo + 1 if mo < 12 else 1
        next_yr = yr if mo < 12 else yr + 1
        # 10e jour ouvré du mois suivant
        d_start  = _date(next_yr, next_mo, 1)
        ouvre    = 0
        d_limit  = d_start
        while ouvre < 10:
            if d_limit.weekday() < 5:
                ouvre += 1
            if ouvre < 10:
                d_limit += timedelta(days=1)
        _write_row(
            "EMEBI (Intrastat)",
            i18n_("Enquête statistique EMEBI {country} (introductions + expéditions, sous réserve de seuil — voir onglet dédié)", country=seller_country),
            f"{yr}-{mo:02d}",
            d_limit,
            "pro.douane.gouv.fr → EMEBI/Intrastat",
            "Art. 7 Règl. UE 2019/2152 — 10e jour ouvré du mois suivant",
            GREEN_FILL,
        )

    # ── 4. Relevé TVA intracom (ESL) ─────────────────────────────────────
    esl_months: set[tuple[int, int]] = set()
    from .models import Scenario as _Scen
    for r in results:
        if r.scenario not in (_Scen.B2B_REVERSE_CHARGE,):
            continue
        d = (r.sale.transaction_date or "")[:10]
        try:
            esl_months.add((int(d[:4]), int(d[5:7])))
        except (ValueError, IndexError):
            pass
    for yr, mo in sorted(esl_months):
        next_mo = mo + 1 if mo < 12 else 1
        next_yr = yr if mo < 12 else yr + 1
        deadline = _date(next_yr, next_mo, 24)
        _write_row(
            i18n_("xl_cal_esl_task"),
            i18n_("xl_cal_esl_desc"),
            f"{yr}-{mo:02d}",
            deadline,
            "impots.gouv.fr → DES (Déclaration Européenne de Services) / ESL",
            "Art. 289 B CGI — même délai que CA3",
            RED_FILL,
        )

    if row == 5:
        ws.append([_wcell(ws, i18n_("xl_cal_no_deadline"), font=Font(italic=True))])

    _width_tracker.apply(ws)


def _parse_fc_transfer(t: dict) -> tuple[str, str, str, str, str, str, int]:
    """Extrait les champs normalisés d'une ligne FC transfer (multi-format).

    Retourne (tx_id, date_str, asin, designation, dep, arr, qty).
    """
    # Transaction ID
    tx_id = (
            t.get("TRANSACTION_EVENT_ID") or t.get("transaction_event_id") or
            t.get("ACTIVITY_TRANSACTION_ID") or t.get("activity_transaction_id") or ""
    )
    # Date
    # BUGFIX : les exports Amazon (transferts FC) fournissent cette date au
    # format "DD-MM-YYYY" (ex: "31-05-2026"), jamais ISO. Le code découpait
    # auparavant cette chaîne comme si elle était déjà "YYYY-MM-DD"
    # (`mois = date_str[:7]`, `annee = mois[:4]` dans _write_intrastat_tab),
    # ce qui produisait des valeurs absurdes ("01-0", "02-0"...) au lieu
    # d'une vraie année, et donc un mauvais regroupement mensuel/annuel des
    # flux Intrastat et un mauvais calcul des dates limites de déclaration
    # (Calendrier Fiscal). `parse_date()` (déjà utilisé par les parsers de
    # ventes pour ce même format) normalise ici vers ISO AVANT tout découpage
    # en aval.
    date_str = _parse_amz_date(
        t.get("TRANSACTION_COMPLETE_DATE") or t.get("transaction_complete_date") or
        t.get("TAX_CALCULATION_DATE") or t.get("tax_calculation_date") or ""
    )
    # ASIN
    asin = (t.get("ASIN") or t.get("asin") or "").strip()
    # Désignation
    designation = (
            t.get("ITEM_DESCRIPTION") or t.get("item_description") or
            t.get("item_name") or ""
    )
    # Pays départ / arrivée
    dep = (
            t.get("DEPARTURE_COUNTRY") or t.get("departure_country") or
            t.get("SALE_DEPART_COUNTRY") or t.get("sale_depart_country") or ""
    ).strip().upper()
    arr = (
            t.get("ARRIVAL_COUNTRY") or t.get("arrival_country") or
            t.get("SALE_ARRIVAL_COUNTRY") or t.get("sale_arrival_country") or ""
    ).strip().upper()
    # Quantité
    raw_qty = t.get("QTY") or t.get("qty") or 1
    try:
        qty = int(float(raw_qty))
    except (ValueError, TypeError):
        qty = 1
        logger.warning(
            "Intrastat/EMEBI : QTY illisible ('%s') pour ASIN=%s, transfert %s→%s "
            "(tx_id=%s) — quantité forcée à 1, seuil EMEBI potentiellement faussé.",
            raw_qty, asin or "?", dep, arr, tx_id or "?",
        )

    return tx_id, date_str, asin, str(designation), dep, arr, qty


def _build_asin_avg_price(results: list) -> dict[str, Decimal]:
    """Calcule le prix de vente HT moyen par ASIN à partir des VatResult de ventes.

    Utilisé comme approximation de la base imposable AIC (valeur d'achat inconnue).
    Seules les ventes avec montant > 0 sont prises en compte (exclut remboursements).
    """
    totals: dict[str, list[Decimal]] = {}
    for r in results:
        asin = getattr(r.sale, "asin", "").strip()
        amt  = r.sale.amount_ht
        if asin and amt > Decimal("0"):
            totals.setdefault(asin, []).append(amt)
    return {
        asin: sum(amounts, Decimal("0")) / Decimal(str(len(amounts)))
        for asin, amounts in totals.items()
        if amounts
    }


def _write_fba_transfers_tab(ws, all_fc_transfers: list) -> None:
    """Onglet Mouvements Stock FBA — détail de chaque transfert."""
    ws.title = i18n_("xl_tab_fba")
    _width_tracker = _ColumnWidthTracker()
    _headers = [
        i18n_("xl_fba_col_tx_id"), i18n_("xl_fba_col_date"), i18n_("xl_fba_col_asin"), i18n_("xl_fba_col_desc"),
        i18n_("xl_fba_col_qty"), i18n_("xl_fba_col_dep"), i18n_("xl_fba_col_arr"), i18n_("xl_fba_col_type"),
    ]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in _headers])
    ws.row_dimensions[1].height = 22
    _width_tracker.observe_row(_headers)

    if not all_fc_transfers:
        ws.append([_wcell(ws, i18n_("xl_fba_none"))])
        _width_tracker.apply(ws)
        return

    for i, t in enumerate(all_fc_transfers, 2):
        tx_id, date_str, asin, designation, dep, arr, qty = _parse_fc_transfer(t)
        tx_type = (t.get("TRANSACTION_TYPE") or t.get("transaction_type") or "FC_TRANSFER").upper()
        _vals = [tx_id, date_str, asin, designation, qty, dep or "—", arr or "—", tx_type]
        ws.append([_wcell(ws, v) for v in _vals])
        ws.row_dimensions[i].height = 18
        _width_tracker.observe_row(_vals)

    _width_tracker.apply(ws)


def _write_fba_aic_tab(
        ws,
        all_fc_transfers: list,
        results: list,
        countries_with_vat: list[str] | None = None,
        display_currency: str = "EUR",
) -> None:
    """Onglet Analyse AIC (Acquisitions Intracommunautaires assimilées).

    Pour chaque flux pays_départ → pays_arrivée où le vendeur est immatriculé
    dans les DEUX pays, calcule une estimation de la TVA AIC à autodéclarer :

        Base AIC estimée  = Σ (qté × prix_vente_moyen_HT_par_ASIN)
        TVA AIC estimée   = Base × taux_standard_pays_arrivée

    ⚠ La base légale AIC est la valeur d'ACHAT (art. 83 directive 2006/112/CE).
    Amazon ne fournissant pas cette donnée, on utilise le prix de vente HT moyen
    comme approximation par excès (prudente, généralement acceptée en pratique).
    Remplacer par le prix d'achat réel si disponible.

    Les flux sans immatriculation dans l'un des deux pays sont listés en
    section "Flux non concernés" pour mémoire.
    """
    from .rates import vat_rate as _vat_rate, STANDARD_VAT_RATES

    ws.title = "Analyse AIC FBA"
    countries_with_vat = [c.upper() for c in (countries_with_vat or [])]
    _width_tracker = _ColumnWidthTracker()

    _fmt_curr = _currency_format(display_currency)
    _conv_date = _date.today()
    def _conv(amount: Decimal) -> float:
        return float(_to_home_currency(amount, display_currency, _conv_date))

    # --- Prix moyen HT par ASIN depuis les ventes ---
    asin_avg = _build_asin_avg_price(results)

    # --- Agrégation par (départ, arrivée, asin) ---
    from collections import defaultdict
    flux_asin: dict[tuple[str, str, str], dict] = defaultdict(lambda: {
        "designation": "", "qty": 0, "nb_transfers": 0,
    })
    flux_summary: dict[tuple[str, str], dict] = defaultdict(lambda: {
        "nb_transfers": 0, "asins": set(),
    })

    for t in all_fc_transfers:
        _, _, asin, designation, dep, arr, qty = _parse_fc_transfer(t)
        if not dep or not arr:
            continue
        key = (dep, arr, asin)
        flux_asin[key]["qty"]          += qty
        flux_asin[key]["nb_transfers"] += 1
        flux_asin[key]["designation"]   = flux_asin[key]["designation"] or designation
        flux_summary[(dep, arr)]["nb_transfers"] += 1
        flux_summary[(dep, arr)]["asins"].add(asin)

    # Séparer flux "à déclarer" (vendeur immatriculé dep ET arr) vs "non concernés"
    flux_actifs   = {k: v for k, v in flux_summary.items()
                     if k[0] in countries_with_vat and k[1] in countries_with_vat}
    flux_inactifs = {k: v for k, v in flux_summary.items()
                     if k not in flux_actifs}

    # ----------------------------------------------------------------
    # En-tête de l'onglet
    # ----------------------------------------------------------------
    ws.append([_wcell(ws, "ANALYSE DES ACQUISITIONS INTRACOMMUNAUTAIRES ASSIMILÉES (FC TRANSFERS)", font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25

    ws.append([_wcell(ws, (
        "⚠ Base AIC estimée = prix de vente HT moyen (Amazon ne fournit pas le prix d'achat). "
        "Approximation par excès — remplacer par le coût d'achat réel si disponible (art. 83 dir. 2006/112/CE)."
    ), font=Font(italic=True, size=10, color="C00000"))])
    ws.row_dimensions[2].height = 30
    ws.append([])

    current_row = 4

    # ----------------------------------------------------------------
    # Section 1 : Flux actifs (immatriculation dans les deux pays)
    # ----------------------------------------------------------------
    ws.append([_wcell(ws, "FLUX AVEC IMMATRICULATION DANS LES DEUX PAYS — AIC À DÉCLARER", font=Font(bold=True, size=11, color="C00000"))])
    current_row += 1

    if not flux_actifs:
        ws.append([_wcell(ws, "Aucun flux ne nécessite de déclaration AIC (immatriculations croisées insuffisantes).")])
        current_row += 2
    else:
        # En-tête détail ASIN
        _headers_detail = [
            "Départ", "Arrivée",
            "ASIN", "Désignation",
            "Qté transférée", f"Prix vente moy. HT ({display_currency})",
            f"Base AIC estimée ({display_currency})", "Taux TVA arrivée (%)",
            f"TVA AIC estimée ({display_currency})", "Statut",
        ]
        ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                          alignment=Alignment(horizontal="center", vertical="center"))
                   for t in _headers_detail])
        ws.row_dimensions[current_row].height = 22
        _width_tracker.observe_row(_headers_detail)
        current_row += 1

        # Regrouper par flux pour les totaux
        flux_totaux: dict[tuple[str, str], dict] = defaultdict(
            lambda: {"base": Decimal("0"), "tva": Decimal("0")}
        )

        for (dep, arr, asin), data in sorted(flux_asin.items()):
            if (dep, arr) not in flux_actifs:
                continue

            qty         = data["qty"]
            designation = data["designation"]
            avg_price   = asin_avg.get(asin, Decimal("0"))
            base_aic    = _round(Decimal(str(qty)) * avg_price)
            taux_arr    = _vat_rate(arr, "STANDARD") if arr in STANDARD_VAT_RATES else Decimal("0")
            tva_aic     = _round(base_aic * taux_arr / Decimal("100"))
            statut      = "✅ Immatriculé" if (dep in countries_with_vat and arr in countries_with_vat) else "🚨 Vérifier"

            flux_totaux[(dep, arr)]["base"] += base_aic
            flux_totaux[(dep, arr)]["tva"]  += tva_aic

            _dep_lbl = f"{_COUNTRY_NAMES_XL.get(dep, dep)} ({dep})"
            _arr_lbl = f"{_COUNTRY_NAMES_XL.get(arr, arr)} ({arr})"
            _desc80 = designation[:80]
            _avg_f, _base_f, _taux_f, _tva_f = float(avg_price), float(base_aic), float(taux_arr), float(tva_aic)
            _vals = [_dep_lbl, _arr_lbl, asin, _desc80, qty, _avg_f, _base_f, _taux_f, _tva_f, statut]
            ws.append([
                _wcell(ws, _dep_lbl), _wcell(ws, _arr_lbl), _wcell(ws, asin), _wcell(ws, _desc80),
                _wcell(ws, qty),
                _wcell(ws, _conv(avg_price), number_format=_fmt_curr),
                _wcell(ws, _conv(base_aic), number_format=_fmt_curr),
                _wcell(ws, float(taux_arr), number_format=_PCT_FORMAT),
                _wcell(ws, _conv(tva_aic), number_format=_fmt_curr, font=_BOLD_FONT),
                _wcell(ws, statut),
            ])
            ws.row_dimensions[current_row].height = 18
            _width_tracker.observe_row(_vals)
            current_row += 1

        # Lignes de sous-total par flux
        current_row += 1
        ws.append([_wcell(ws, "SOUS-TOTAUX PAR FLUX", font=Font(bold=True, size=10))])
        current_row += 1
        _headers_sub = [
            "Flux (Départ → Arrivée)", "Nb transferts", "Nb ASIN",
            f"Base AIC totale estimée ({display_currency})", f"TVA AIC totale estimée ({display_currency})",
            "Référence légale", "Action requise",
        ]
        ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                          alignment=Alignment(horizontal="center", vertical="center"))
                   for t in _headers_sub])
        ws.row_dimensions[current_row].height = 22
        _width_tracker.observe_row(_headers_sub)
        current_row += 1

        for (dep, arr) in sorted(flux_actifs):
            nb_t  = flux_actifs[(dep, arr)]["nb_transfers"]
            nb_a  = len(flux_actifs[(dep, arr)]["asins"])
            base  = flux_totaux[(dep, arr)]["base"]
            tva   = flux_totaux[(dep, arr)]["tva"]
            ref   = f"AIC art. 17 dir. 2006/112/CE — déclarer en TVA {arr}"
            action = f"Inclure {float(tva):,.2f} € en TVA {arr} (autodéclaration)"
            _flow_lbl = f"{_COUNTRY_NAMES_XL.get(dep, dep)} → {_COUNTRY_NAMES_XL.get(arr, arr)}"
            _base_f, _tva_f = float(base), float(tva)
            _vals_sub = [_flow_lbl, nb_t, nb_a, _base_f, _tva_f, ref, action]
            ws.append([
                _wcell(ws, _flow_lbl), _wcell(ws, nb_t), _wcell(ws, nb_a),
                _wcell(ws, _conv(base), number_format=_fmt_curr, font=_BOLD_FONT),
                _wcell(ws, _conv(tva), number_format=_fmt_curr, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL),
                _wcell(ws, ref), _wcell(ws, action),
            ])
            ws.row_dimensions[current_row].height = 20
            _width_tracker.observe_row(_vals_sub)
            current_row += 1

    current_row += 2

    # ----------------------------------------------------------------
    # Section 2 : Flux sans double immatriculation (pour mémoire)
    # ----------------------------------------------------------------
    ws.append([])
    ws.append([_wcell(ws, "FLUX SANS IMMATRICULATION CROISÉE — POUR MÉMOIRE (Amazon gère)", font=Font(bold=True, size=11, color="808080"))])
    current_row += 1

    if not flux_inactifs:
        ws.append([_wcell(ws, "—")])
        current_row += 1
    else:
        _headers_inact = [
            "Départ", "Arrivée", "Nb transferts", "Nb ASIN distincts",
            "Immat. départ", "Immat. arrivée", "Observation",
        ]
        ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE,
                          fill=PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"),
                          alignment=Alignment(horizontal="center", vertical="center"))
                   for t in _headers_inact])
        ws.row_dimensions[current_row].height = 22
        _width_tracker.observe_row(_headers_inact)
        current_row += 1

        for (dep, arr) in sorted(flux_inactifs):
            nb_t = flux_inactifs[(dep, arr)]["nb_transfers"]
            nb_a = len(flux_inactifs[(dep, arr)]["asins"])
            imm_dep = "✅" if dep in countries_with_vat else "—"
            imm_arr = "✅" if arr in countries_with_vat else "—"
            if dep not in countries_with_vat and arr not in countries_with_vat:
                obs = "Aucune immatriculation — Amazon gère l'AIC"
            elif dep in countries_with_vat:
                obs = f"LIC à déclarer côté {dep} (case exonérations)"
            else:
                obs = f"Vérifier immatriculation {arr}"
            _dep_lbl2 = f"{_COUNTRY_NAMES_XL.get(dep, dep)} ({dep})"
            _arr_lbl2 = f"{_COUNTRY_NAMES_XL.get(arr, arr)} ({arr})"
            _vals_inact = [_dep_lbl2, _arr_lbl2, nb_t, nb_a, imm_dep, imm_arr, obs]
            ws.append([_wcell(ws, v) for v in _vals_inact])
            ws.row_dimensions[current_row].height = 18
            _width_tracker.observe_row(_vals_inact)
            current_row += 1

    _width_tracker.apply(ws)


def _month_label(month_key: str) -> str:
    """Formate une clé "YYYY-MM" en libellé colonne lisible "MM/YYYY"."""
    y, _sep, m = month_key.partition("-")
    return f"{m}/{y}" if m else month_key


def _write_section_group_row(ws, month_start_col: int, n_months: int, total_start_col: int, n_total_cols: int, fill) -> list:
    """Construit une ligne de regroupement au-dessus des en-têtes de colonnes :
    un bandeau coloré sur les colonnes mois ("Détail mensuel (net)") et un
    bandeau coloré sur les colonnes de total période ("Total période").
    Ne fait rien pour la partie mensuelle si n_months == 0.

    Fusion de cellules abandonnée (incompatible write_only, voir décision du
    chantier) : pour préserver le rendu visuel du bandeau coloré (pas
    seulement un débordement de texte sur fond blanc), toutes les cellules
    du groupe reçoivent le même remplissage — seul le libellé est posé sur
    la première cellule de chaque groupe.
    """
    last_col = total_start_col + n_total_cols - 1
    row_cells = [_wcell(ws, None) for _ in range(last_col)]  # colonnes 1..last_col, vides par défaut

    if n_months:
        first, last = month_start_col, month_start_col + n_months - 1
        for col in range(first, last + 1):
            # On utilise i18n_ (aliasé ci-dessous) pour éviter le shadowing par _ dans les boucles
            from .i18n import _ as i18n_
            c = _wcell(ws, i18n_("xl_monthly_section_label") if col == first else None,
                       font=_HEADER_FONT_WHITE, fill=fill,
                       alignment=Alignment(horizontal="center", vertical="center"))
            row_cells[col - 1] = c

    first, last = total_start_col, total_start_col + n_total_cols - 1
    for col in range(first, last + 1):
        from .i18n import _ as i18n_
        c = _wcell(ws, i18n_("xl_period_section_label") if col == first else None,
                   font=_HEADER_FONT_WHITE, fill=fill,
                   alignment=Alignment(horizontal="center", vertical="center"))
        row_cells[col - 1] = c

    return row_cells


def _write_oss_tab(ws, summary: ReportSummary, display_currency: str = "EUR",
                   results: list | None = None, refund_results: list | None = None,
                   period: str = "") -> None:
    """Onglet OSS détaillé : mois par mois (net) puis Brut / Remboursements / Net
    (total période) par pays de destination."""
    ws.title = i18n_("xl_tab_oss")
    _width_tracker = _ColumnWidthTracker()

    ws.append([_wcell(ws, i18n_("xl_oss_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25

    _fmt_curr = _currency_format(display_currency)
    _conv_date = _date.today()
    def _conv(amount: Decimal) -> float:
        return float(_to_home_currency(amount, display_currency, _conv_date))

    if display_currency != "EUR":
        ws.append([_wcell(ws, i18n_("xl_recap_currency_note", currency=display_currency, date=_conv_date.isoformat()),
                          font=Font(italic=True, size=9, color="7f7f7f"))])
        ws.row_dimensions[2].height = 16
    else:
        ws.append([])  # ligne 2 volontairement vide (comportement d'origine)

    _z = Decimal("0.00")
    all_countries = sorted(
        set(summary.oss_by_country) | set(getattr(summary, "refund_oss_by_country", {}))
    )
    by_country_month = getattr(summary, "oss_by_country_month", {}) or {}
    months = sorted({m for per_country in by_country_month.values() for m in per_country})

    # Totaux Brut/Remb par pays d'arrivée reconvertis au taux BCE de clôture
    # de période (art. 5 bis Règl. UE 2020/194) — même méthode que le
    # dashboard (declarations.py). `summary.oss_by_country` seul reste figé
    # au taux du jour de vente : pour les pays facturés en devise étrangère
    # (ex. Suède/SEK), ça produit un total différent de celui affiché au
    # tableau de bord. On ne recalcule que si `results` a été fourni ;
    # sinon on retombe sur le comportement historique (summary seul).
    _period_tva_vente: Dict[str, Decimal] = {}
    _period_tva_remb: Dict[str, Decimal] = {}
    if results is not None:
        _agg = aggregate_oss_results(list(results) + list(refund_results or []), period=period)
        for _departure, _by_arrival in _agg.items():
            for _arrival, _by_rate in _by_arrival.items():
                for _bucket in _by_rate.values():
                    _period_tva_vente[_arrival] = _period_tva_vente.get(_arrival, _z) + _bucket["tva_vente"]
                    _period_tva_remb[_arrival] = _period_tva_remb.get(_arrival, _z) + _bucket["tva_remb"]
        all_countries = sorted(set(all_countries) | set(_period_tva_vente) | set(_period_tva_remb))

    # Colonnes : Pays, Code, [mois...] (net seul), Brut, Remboursements, Net (total période)
    month_start_col = 3
    total_start_col = month_start_col + len(months)

    header_row = 4
    _group_cells = _write_section_group_row(ws, month_start_col, len(months), total_start_col, 3, fill=_BLUE_HEADER_FILL)
    ws.append(_group_cells)
    ws.row_dimensions[3].height = 18

    headers = [i18n_("xl_oss_col_country"), i18n_("xl_oss_col_code")]
    headers += [_month_label(m) for m in months]
    headers += [i18n_("xl_oss_col_vat_gross"), i18n_("xl_oss_col_vat_refunds"), i18n_("xl_oss_col_vat_net")]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in headers])
    ws.row_dimensions[header_row].height = 22
    _width_tracker.observe_row(headers)

    row = header_row + 1
    for country in all_countries:
        if results is not None:
            brut   = _period_tva_vente.get(country, _z)
            refund = _period_tva_remb.get(country, _z)
        else:
            brut   = summary.oss_by_country.get(country, _z)
            refund = summary.refund_oss_by_country.get(country, _z) if getattr(summary, "refund_oss_by_country", None) else _z
        net    = brut + refund  # noqa: F841 (conservé pour parité de lecture avec l'original)

        month_values = by_country_month.get(country, {})
        col_brut, col_ref, col_net = total_start_col, total_start_col + 1, total_start_col + 2
        letter_brut, letter_ref, letter_net = get_column_letter(col_brut), get_column_letter(col_ref), get_column_letter(col_net)

        _vals = [_get_country_name(country), country]
        _row_cells = [_wcell(ws, _get_country_name(country)), _wcell(ws, country)]
        for m in months:
            v = _conv(month_values.get(m, _z))
            _vals.append(v)
            _row_cells.append(_wcell(ws, v, number_format=_fmt_curr))

        _vals += [_conv(brut), _conv(refund), f"={letter_brut}{row}+{letter_ref}{row}"]
        _row_cells.append(_wcell(ws, _conv(brut), number_format=_fmt_curr))
        _row_cells.append(_wcell(ws, _conv(refund), number_format=_fmt_curr))
        _row_cells.append(_wcell(ws, f"={letter_brut}{row}+{letter_ref}{row}",
                                 number_format=_fmt_curr, font=_BOLD_FONT, fill=_LIGHT_GRAY_FILL))
        # Excel recalculates this correctly on open, but we help it
        # by ensuring letters match the displayed screenshot bug (G5 = Brut, H5 = Refund).
        # Wait, in the code letter_brut is col_brut. col_brut = total_start_col.
        # If months is 3 (Jan, Feb, Mar), month_start_col=3 (C).
        # C, D, E are months. F is Brut. G is Refund. H is Net.
        # Screenshot shows F=Brut, G=Refund, H=Net. This matches col_brut=6 (F).

        ws.append(_row_cells)
        ws.row_dimensions[row].height = 18
        _width_tracker.observe_row(_vals)
        row += 1

    # Ligne de total (une ligne blanche d'écart avant, comme dans l'original)
    col_brut, col_ref, col_net = total_start_col, total_start_col + 1, total_start_col + 2
    letter_brut, letter_ref, letter_net = get_column_letter(col_brut), get_column_letter(col_ref), get_column_letter(col_net)
    ws.append([])
    row += 1
    _total_row_cells = [_wcell(ws, i18n_("xl_oss_total"), font=_BOLD_FONT)]
    _total_row_cells.append(_wcell(ws, None))  # colonne "Code", vide
    for i in range(len(months)):
        col = month_start_col + i
        letter = get_column_letter(col)
        _total_row_cells.append(_wcell(ws, f"=SUM({letter}{header_row+1}:{letter}{row-2})",
                                       number_format=_fmt_curr, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL))
    for formula in [
        f"=SUM({letter_brut}{header_row+1}:{letter_brut}{row-2})",
        f"=SUM({letter_ref}{header_row+1}:{letter_ref}{row-2})",
        f"={get_column_letter(col_brut)}{row}+{get_column_letter(col_ref)}{row}",
    ]:
        _total_row_cells.append(_wcell(ws, formula, number_format=_fmt_curr, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL))
    ws.append(_total_row_cells)
    ws.row_dimensions[row].height = 20

    _width_tracker.apply(ws)





def _write_local_tab(ws, summary: ReportSummary, countries_with_vat: list | None = None, seller_country: str = "FR", display_currency: str = "EUR") -> None:
    """Onglet TVA locale par pays (immatriculation locale hors OSS) : mois par
    mois (net) puis Brut / Remboursements / Net (total période) et statut."""
    ws.title = i18n_("xl_tab_local")
    countries_with_vat = {c.upper() for c in (countries_with_vat or [])}
    # Le pays d'origine est toujours considéré comme immatriculé
    countries_with_vat.add(seller_country.upper())

    ws.append([_wcell(ws, i18n_("xl_local_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25
    _width_tracker = _ColumnWidthTracker()

    _fmt_curr = _currency_format(display_currency)
    _conv_date = _date.today()
    def _conv(amount: Decimal) -> float:
        return float(_to_home_currency(amount, display_currency, _conv_date))

    if display_currency != "EUR":
        ws.append([_wcell(ws, i18n_("xl_recap_currency_note", currency=display_currency, date=_conv_date.isoformat()),
                          font=Font(italic=True, size=9, color="7f7f7f"))])
        ws.row_dimensions[2].height = 16
    else:
        ws.append([])
    _z = Decimal("0.00")
    local = dict(summary.local_by_country or {})
    refund_local = dict(getattr(summary, "refund_local_by_country", {}) or {})

    # Ajouter le pays d'origine s'il y a de la TVA domestique
    if summary.fr_domestic_vat or summary.refund_fr_domestic_vat:
        sc = seller_country.upper()
        local[sc] = local.get(sc, _z) + summary.fr_domestic_vat
        refund_local[sc] = refund_local.get(sc, _z) + summary.refund_fr_domestic_vat

    all_countries = sorted(set(local) | set(refund_local))
    unregistered = [c for c in all_countries if c not in countries_with_vat]

    by_country_month = dict(getattr(summary, "local_by_country_month", {}) or {})
    # Injecter les données mensuelles du pays d'origine
    if getattr(summary, "fr_domestic_by_month", None):
        sc = seller_country.upper()
        existing = by_country_month.get(sc, {})
        for m, val in summary.fr_domestic_by_month.items():
            existing[m] = existing.get(m, _z) + val
        by_country_month[sc] = existing

    months = sorted({m for per_country in by_country_month.values() for m in per_country})

    # BUGFIX (#VALEUR! en colonne "TVA Nette") : cet onglet insère UNE ligne
    # de plus que l'onglet OSS avant les en-têtes (l'avertissement "pays non
    # immatriculé" OU une ligne vide de remplacement, voir juste en-dessous),
    # ce qui décale la ligne d'en-têtes réelle à la ligne 5 (et non 4 comme
    # dans _write_oss_tab, qui n'a pas cette ligne supplémentaire). `header_row`
    # restait à 4 alors que la ligne physique des en-têtes est 5 : chaque
    # formule `=E{row}+F{row}` référençait donc la ligne du DESSUS (en-tête ou
    # pays précédent) au lieu de sa propre ligne, d'où des totaux décalés
    # d'une ligne et un #VALEUR! sur la dernière ligne (qui se retrouvait à
    # additionner la ligne "TOTAL LOCAL", du texte).
    header_row = 5
    if unregistered:
        ws.append([_wcell(ws, i18n_("xl_local_unregistered_warning", countries=", ".join(unregistered)),
                          font=_ALERT_FONT, fill=_ALERT_FILL)])
        ws.row_dimensions[3].height = 18
    else:
        ws.append([])

    # Colonnes : Pays, Code, [mois...] (net seul), Brut, Remboursements, Net (total période), Statut
    month_start_col = 3
    total_start_col = month_start_col + len(months)
    col_brut, col_ref, col_net, col_status = total_start_col, total_start_col + 1, total_start_col + 2, total_start_col + 3
    letter_brut, letter_ref = get_column_letter(col_brut), get_column_letter(col_ref)

    _group_cells = _write_section_group_row(ws, month_start_col, len(months), total_start_col, header_row - 1, fill=_ORANGE_HEADER_FILL)
    ws.append(_group_cells)
    ws.row_dimensions[header_row - 1].height = 18

    headers = [i18n_("xl_local_col_country"), i18n_("xl_local_col_code")]
    headers += [_month_label(m) for m in months]
    headers += [i18n_("xl_local_col_vat_due"), i18n_("xl_local_col_vat_refunds"), i18n_("xl_local_col_vat_net"), i18n_("xl_local_col_status")]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in headers])
    ws.row_dimensions[header_row].height = 22
    _width_tracker.observe_row(headers)

    row = header_row + 1
    for country in all_countries:
        brut   = local.get(country, _z)
        refund = refund_local.get(country, _z)
        is_registered = country in countries_with_vat

        month_values = by_country_month.get(country, {})
        _vals = [_get_country_name(country), country]
        _row_cells = [_wcell(ws, _get_country_name(country)), _wcell(ws, country)]
        for m in months:
            v = _conv(month_values.get(m, _z))
            _vals.append(v)
            _row_cells.append(_wcell(ws, v, number_format=_fmt_curr))

        _row_cells.append(_wcell(ws, _conv(brut), number_format=_fmt_curr))
        _row_cells.append(_wcell(ws, _conv(refund), number_format=_fmt_curr))
        _row_cells.append(_wcell(ws, f"={letter_brut}{row}+{letter_ref}{row}",
                                 number_format=_fmt_curr, font=_BOLD_FONT, fill=_LIGHT_GRAY_FILL))
        # Excel recalculates this correctly on open, but we help it
        # by ensuring letters match the displayed screenshot bug (G5 = Brut, H5 = Refund).
        # Wait, in the code letter_brut is col_brut. col_brut = total_start_col.
        # If months is 3 (Jan, Feb, Mar), month_start_col=3 (C).
        # C, D, E are months. F is Brut. G is Refund. H is Net.
        # Screenshot shows F=Brut, G=Refund, H=Net. This matches col_brut=6 (F).
        _status_val = i18n_("xl_local_status_registered") if is_registered else i18n_("xl_local_status_unconfirmed")
        _row_cells.append(_wcell(ws, _status_val,
                                 font=_ALERT_FONT if not is_registered else None,
                                 fill=_ALERT_FILL if not is_registered else None))
        _vals += [_conv(brut), _conv(refund), _status_val]

        ws.append(_row_cells)
        ws.row_dimensions[row].height = 18
        _width_tracker.observe_row(_vals)
        row += 1

    # Total
    ws.append([])
    row += 1
    _total_cells = [_wcell(ws, i18n_("xl_local_total"), font=_BOLD_FONT), _wcell(ws, None)]
    for i in range(len(months)):
        col = month_start_col + i
        letter = get_column_letter(col)
        _total_cells.append(_wcell(ws, f"=SUM({letter}{header_row+1}:{letter}{row-2})",
                                   number_format=_fmt_curr, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL))

    # We re-calculate letters for the total row to be absolutely safe
    l_brut = get_column_letter(col_brut)
    l_ref = get_column_letter(col_ref)

    for formula in [
        f"=SUM({l_brut}{header_row+1}:{l_brut}{row-2})",
        f"=SUM({l_ref}{header_row+1}:{l_ref}{row-2})",
        f"={l_brut}{row}+{l_ref}{row}",
    ]:
        _total_cells.append(_wcell(ws, formula, number_format=_fmt_curr, font=_HEADER_FONT_WHITE, fill=_ORANGE_HEADER_FILL))
    _total_cells.append(_wcell(ws, None))  # colonne Statut, vide sur la ligne de total
    ws.append(_total_cells)
    ws.row_dimensions[row].height = 20

    _width_tracker.apply(ws)


def _write_invoice_creditnote_tab(ws, invoice_credit_notes: list) -> None:
    """Onglet INVOICE / CREDIT_NOTE."""
    ws.title = i18n_("xl_tab_invoice_cn")
    _width_tracker = _ColumnWidthTracker()

    ws.append([_wcell(ws, i18n_("xl_inv_cn_title"), font=_TITLE_FONT)])
    ws.row_dimensions[1].height = 25
    ws.append([_wcell(ws, i18n_("xl_inv_cn_help"))])
    ws.row_dimensions[2].height = 18
    ws.append([])

    headers = [i18n_("xl_inv_cn_col_type"), i18n_("xl_inv_cn_col_date"), i18n_("xl_inv_cn_col_market"), i18n_("xl_inv_cn_col_program"), i18n_("xl_inv_cn_col_ref"), i18n_("xl_inv_cn_col_ht"), i18n_("xl_inv_cn_col_vat"), i18n_("xl_inv_cn_col_currency")]
    ws.append([_wcell(ws, t, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL,
                      alignment=Alignment(horizontal="center", vertical="center"))
               for t in headers])
    ws.row_dimensions[4].height = 22
    _width_tracker.observe_row(headers)

    if not invoice_credit_notes:
        ws.append([_wcell(ws, i18n_("xl_inv_cn_none"))])
        _width_tracker.apply(ws)
        return

    row = 5
    total_ht = Decimal("0.00")
    total_vat = Decimal("0.00")
    for entry in invoice_credit_notes:
        amount_ht = entry.get("amount_ht", Decimal("0")) or Decimal("0")
        vat_amount = entry.get("vat_amount", Decimal("0")) or Decimal("0")
        _vals = [
            entry.get("kind", ""), entry.get("date", ""), entry.get("marketplace", ""),
            entry.get("program_type", ""), entry.get("reference", ""),
            float(amount_ht), float(vat_amount), entry.get("currency", "EUR"),
        ]
        ws.append([
            _wcell(ws, _vals[0]), _wcell(ws, _vals[1]), _wcell(ws, _vals[2]), _wcell(ws, _vals[3]), _wcell(ws, _vals[4]),
            _wcell(ws, _vals[5], number_format=_EUR_FORMAT),
            _wcell(ws, _vals[6], number_format=_EUR_FORMAT),
            _wcell(ws, _vals[7]),
        ])

        total_ht += amount_ht
        total_vat += vat_amount
        ws.row_dimensions[row].height = 18
        _width_tracker.observe_row(_vals)
        row += 1

    ws.append([])
    row += 1
    ws.append([
        _wcell(ws, i18n_("xl_total"), font=_BOLD_FONT), _wcell(ws, None), _wcell(ws, None), _wcell(ws, None), _wcell(ws, None),
        _wcell(ws, float(_round(total_ht)), number_format=_EUR_FORMAT, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL),
        _wcell(ws, float(_round(total_vat)), number_format=_EUR_FORMAT, font=_HEADER_FONT_WHITE, fill=_BLUE_HEADER_FILL),
        _wcell(ws, None),
    ])
    ws.row_dimensions[row].height = 20

    _width_tracker.apply(ws)


def export_xlsx(
        results: List[VatResult],
        output_path: str | Path,
        scope_id: str,
        summary: ReportSummary | None = None,
        refund_results: List[VatResult] | None = None,
        all_fc_transfers: list | None = None,
        vies_affected_sale_ids: set | None = None,
        vies_summary=None,
        countries_with_vat: list[str] | None = None,
        period: str = "",
        seller_country: str = "FR",
        display_currency: str | None = None,
        invoice_credit_notes: list | None = None,
) -> Path:
    """Genere le fichier Excel complet avec tous les onglets.

    Args:
        scope_id: portée de cache VIES du compte appelant (voir
                  vies.resolve_scope_id) — transmise à l'onglet Historique
                  VIES pour n'afficher que les vérifications de ce compte.
        display_currency: devise d'affichage choisie pour le rapport (ex: PLN).
                          Si None, utilise la devise du pays d'origine.
    """

    if summary is None:
        summary = build_report(results)

    # Calcul des Hash Totals (Contrôle d'intégrité technique)
    all_rows = results + (refund_results or [])
    hash_totals = {
        "count": len(all_rows),
        "abs_ht": sum((abs(r.sale.amount_ht) for r in all_rows), Decimal("0.00")),
        "vat": sum((abs(r.vat_amount) for r in all_rows), Decimal("0.00")),
        "id_hash": 0,
        "net_ht_check": sum((r.sale.amount_ht for r in all_rows), Decimal("0.00")),
    }
    for r in all_rows:
        # Somme numérique des IDs pour détecter les doublons ou omissions
        raw_id = re.sub(r"\D", "", str(r.sale.sale_id))
        if raw_id:
            # On prend les 6 derniers chiffres pour plus de précision
            hash_totals["id_hash"] += int(raw_id[-6:])

    # write_only=True : toutes les feuilles ci-dessous sont désormais écrites en
    # mode séquentiel (ws.append), ce qui élimine le gonflement mémoire d'openpyxl
    # (mesuré à ~625 Mo pour un export de 20k lignes en mode normal, contre un
    # fichier final de ~2 Mo — voir chantier RAM). Une feuille write_only ne peut
    # plus être relue (pas de wb.active, pas de ws.cell(row=,col=), pas de
    # merge_cells) : chaque fonction _write_*_tab a été adaptée en conséquence.
    wb = Workbook(write_only=True)
    # openpyxl n'écrit jamais de valeur mise en cache pour les cellules-formule
    # (seulement la chaîne "=..."). Excel les recalcule normalement à
    # l'ouverture, mais dans certains cas (Mode protégé après téléchargement,
    # aperçu par certains lecteurs tiers) le classeur peut s'afficher sans
    # recalcul tant que l'utilisateur n'a pas quitté ce mode. `fullCalcOnLoad`
    # force Excel à recalculer TOUT le classeur dès l'ouverture (dès que
    # l'édition/le calcul est autorisé), plutôt que d'attendre une modification
    # manuelle d'une cellule pour déclencher le calcul.
    wb.calculation.fullCalcOnLoad = True

    # 1. Page de synthèse
    ws_recap = _SequentialSheetWriter(wb.create_sheet())
    _write_recap(ws_recap, summary, hash_totals=hash_totals, seller_country=seller_country,
                 display_currency=display_currency, results=results, refund_results=refund_results,
                 period=period)
    ws_recap.finalize()

    # 2. Séparation ventes / remboursements
    # Si refund_results est passé explicitement par app.py (cas normal), on fait
    # confiance à cette séparation : results = ventes uniquement, refund_results = avoirs.
    # On filtre quand même results pour écarter d'éventuels résidus négatifs qui
    # auraient glissé (défense en profondeur), mais on n'ajoute PAS refund_results
    # une deuxième fois s'il est déjà fourni — ce serait un doublon.
    sales_results = []
    refunds_from_results = []  # avoirs détectés dans results (cas mixte ou CLI sans séparation)

    for r in results:
        tx_type  = str(getattr(r.sale, "transaction_type", "")).upper()
        sale_id  = str(getattr(r.sale, "sale_id", "")).upper()
        is_refund = getattr(r.sale, "is_refund", False)

        if tx_type == "REFUND" or is_refund or r.sale.amount_ht < 0 or "REFUND" in sale_id:
            refunds_from_results.append(r)
        else:
            sales_results.append(r)

    # Construire la liste finale des remboursements sans doublon :
    # - Si refund_results fourni explicitement → on l'utilise en priorité et on
    #   ignore refunds_from_results (ils sont déjà dans refund_results).
    # - Sinon (CLI, appel direct) → on utilise ce qu'on a extrait de results.
    if refund_results:
        refunds_results_to_write = list(refund_results)
    else:
        refunds_results_to_write = refunds_from_results

    _currency = display_currency or _home_currency(seller_country)

    # 4. Onglet Détail Ventes
    ws_sales = _SequentialSheetWriter(wb.create_sheet())
    _write_details_tab(ws_sales, "Detail ventes", sales_results, is_refund_tab=False, display_currency=_currency)
    ws_sales.finalize()

    # 5. Onglet Détail Remboursements
    ws_refunds = _SequentialSheetWriter(wb.create_sheet())
    _write_details_tab(ws_refunds, "Detail remboursements", refunds_results_to_write, is_refund_tab=True, display_currency=_currency)
    ws_refunds.finalize()

    # 6. Onglet OSS détaillé par pays
    if summary.oss_by_country or getattr(summary, "refund_oss_by_country", None):
        ws_oss = _SequentialSheetWriter(wb.create_sheet())
        _write_oss_tab(ws_oss, summary, display_currency=_currency,
                       results=results, refund_results=refund_results, period=period)
        ws_oss.finalize()

    # 7. Onglet TVA locale par pays
    if (summary.local_by_country or getattr(summary, "refund_local_by_country", None) or
            summary.fr_domestic_vat or summary.refund_fr_domestic_vat):
        ws_local = _SequentialSheetWriter(wb.create_sheet())
        _write_local_tab(ws_local, summary, countries_with_vat, seller_country=seller_country, display_currency=_currency)
        ws_local.finalize()

    # 8. Onglet Audit Ecarts Amazon
    ws_audit = _SequentialSheetWriter(wb.create_sheet("Audit Ecarts Amazon"))
    _write_audit_tab(ws_audit, results, vies_affected_sale_ids, vies_summary=vies_summary, display_currency=_currency)
    ws_audit.finalize()

    # 8bis. Onglet Historique VIES (piste d'audit — preuve de bonne foi)
    ws_vies_hist = _SequentialSheetWriter(wb.create_sheet("Historique VIES"))
    _write_vies_history_tab(ws_vies_hist, results + (refund_results or []), scope_id)
    ws_vies_hist.finalize()

    # 9. Onglet Analyse AIC FBA (synthèse fiscale des transferts)
    ws_aic = _SequentialSheetWriter(wb.create_sheet("Analyse AIC FBA"))
    _write_fba_aic_tab(ws_aic, all_fc_transfers or [], results, countries_with_vat, display_currency=_currency)
    ws_aic.finalize()

    # 10. Onglet Transferts FBA Détail (liste brute)
    ws_fba = _SequentialSheetWriter(wb.create_sheet("Transferts FBA Détail"))
    _write_fba_transfers_tab(ws_fba, all_fc_transfers or [])
    ws_fba.finalize()

    # 11. Onglet Intrastat / DEB (aide au remplissage)
    ws_intrastat = _SequentialSheetWriter(wb.create_sheet("Intrastat (EMEBI)"))
    _write_intrastat_tab(ws_intrastat, all_fc_transfers or [], results, seller_country=seller_country, display_currency=_currency)
    ws_intrastat.finalize()

    # 11bis. Onglet INVOICE / CREDIT_NOTE (écritures Amazon hors ventes)
    if invoice_credit_notes:
        ws_inv_cn = _SequentialSheetWriter(wb.create_sheet())
        _write_invoice_creditnote_tab(ws_inv_cn, invoice_credit_notes)
        ws_inv_cn.finalize()

    # 12. Onglet Calendrier fiscal (échéances déduites des données)
    ws_cal = _SequentialSheetWriter(wb.create_sheet("Calendrier Fiscal"))
    _write_calendar_tab(
        ws_cal, results, all_fc_transfers or [],
        period=period, seller_country=seller_country,
                         )
    ws_cal.finalize()

    # 13. Sauvegarde sur disque
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(p))
    return p