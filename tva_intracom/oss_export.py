"""Export des états récapitulatifs fiscaux : OSS et B2B reverse charge.

Génère :
- Un fichier Excel (.xlsx) avec 3 onglets :
    1. OSS_Résumé      : TVA due par pays de destination (portail OSS URSSAF)
    2. OSS_Détail      : ligne par ligne des ventes OSS
    3. B2B_Recap       : livraisons intracommunautaires B2B avec numéros TVA
- Deux fichiers CSV :
    1. oss_urssaf.csv  : format portail OSS URSSAF (pays, base HT, taux, TVA)
    2. b2b_recap.csv   : état récapitulatif B2B avec numéros TVA acheteurs

Usage:
    from tva_intracom.oss_export import build_oss_export
    xlsx_path, oss_csv, b2b_csv = build_oss_export(results, period="2024-T1")
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date as _date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from .ecb_rates import convert_to_currency_for_oss, get_oss_rate_date, prefetch_rates
from .i18n import _, country_label
from .models import Scenario, VatResult
from .rates import fiscal_equivalent_country

_CENT = Decimal("0.01")
_ZERO = Decimal("0.00")

# ---------------------------------------------------------------------------
# Agrégation partagée — utilisée par oss_export.py ET oss_xml.py
# ---------------------------------------------------------------------------

# Type : départ → arrivée → taux → {ht, tva, nb}
OssAggType = dict  # dict[str, dict[str, dict[Decimal, dict[str, Decimal | int]]]]


def convert_ht_tva_for_oss_period(res: VatResult, period: str) -> tuple[Decimal, Decimal]:
    """Retourne (ht, tva) d'un VatResult OSS, reconverti au besoin au taux BCE
    de clôture de la période déclarée (Règl. UE 2020/194, art. 5 bis).

    Si `period` est vide/non reconnu, ou si la vente est déjà dans la devise cible, on
    retombe sur `res.sale.amount_ht` / `res.vat_amount` tels quels.
    """
    ht  = res.sale.amount_ht
    tva = res.vat_amount
    
    # BUGFIX CRITIQUE : la déclaration OSS est légalement due en EUR (Règl. UE
    # 2020/194, art. 5 bis) — cette fonction alimente aussi bien le XML OSS
    # officiel (oss_xml.py) que l'export Excel/CSV URSSAF. Utiliser la devise
    # du pays d'origine (home_country) ici corromprait la déclaration légale
    # elle-même, pas seulement un rapport d'affichage. La conversion vers une
    # devise d'affichage locale se fait uniquement en couche présentation
    # (voir tva_intracom/ui/formatting.py, report.py, excel_report.py).
    target_currency = "EUR"

    if period and res.sale.original_currency and res.sale.original_currency != target_currency:
        try:
            tx_date = _date.fromisoformat((res.sale.transaction_date or "")[:10])
        except ValueError:
            tx_date = _date.today()
        sign = Decimal("-1") if ht < 0 else Decimal("1")
        try:
            new_ht_abs, _rate_used, _src = convert_to_currency_for_oss(
                abs(res.sale.original_amount),
                res.sale.original_currency,
                target_currency,
                period,
                tx_date,
                fallback_rate=res.sale.exchange_rate or None,
            )
            ht = sign * new_ht_abs
            tva = (ht * (res.vat_rate / Decimal("100"))).quantize(_CENT, rounding=ROUND_HALF_UP)
        except ValueError:
            # BCE indisponible et pas de fallback : on garde le montant déjà
            # converti au taux du jour de la vente plutôt que de bloquer
            # toute la génération du rapport.
            #
            # C'est ICI (et pas dans convert_to_currency_for_oss elle-même)
            # que vit le repli "taux du jour de la transaction" documenté
            # dans la docstring de convert_to_currency_for_oss (voir
            # ecb_rates.py) — voir sa note "PRÉCISION (audit du 2026-08-19)"
            # pour le détail de ce partage de responsabilité entre les deux
            # fonctions.
            pass

    return ht, tva


def aggregate_oss_results(results: list[VatResult], period: str = "") -> OssAggType:
    """Agrège les VatResult OSS_B2C par pays de départ puis pays d'arrivée.

    ⚠️ CORRECTIF 2026-08-09 : ne traite PLUS Scenario.IOSS_DIRECT (voir
    aggregate_ioss_results() ci-dessous, ajoutée séparément). OSS et IOSS
    sont deux régimes distincts avec des périodicités différentes
    (trimestrielle pour l'OSS, mensuelle pour l'IOSS — art. 369a-k vs.
    art. 369l-x dir. 2006/112/CE) et des numéros d'identification distincts.
    Les mélanger dans une même agrégation produisait un double problème :
    un total OSS trimestriel incluant à tort des montants IOSS mensuels
    dans l'Excel/CSV URSSAF, ET une disparition SILENCIEUSE des montants
    IOSS dans le XML officiel (oss_xml.py filtre les pays de départ hors UE,
    qui est systématiquement le cas pour l'IOSS — import depuis un pays
    tiers). Voir aggregate_ioss_results() + build_ioss_excel/build_ioss_csv
    pour l'export IOSS désormais séparé.

    Structure retournée (utilisée par oss_xml.py pour le XML officiel et
    par oss_export.py pour l'Excel/CSV URSSAF) :

        {
          "FR": {
            "DE": {
              Decimal("19"): {
                  "ht": Decimal(...), "tva": Decimal(...),        # net (vente+avoir)
                  "ht_vente": Decimal(...), "tva_vente": Decimal(...),   # ventes seules (brut)
                  "ht_remb":  Decimal(...), "tva_remb":  Decimal(...),   # avoirs seuls (négatif)
                  "nb": int,
              },
              ...
            },
          },
          "DE": { ... },
        }

    Les clés "ht"/"tva" (net) sont historiques — c'est ce que consomme
    oss_xml.py et find_oss_negative_buckets(). Les clés "*_vente"/"*_remb"
    sont ajoutées pour permettre un affichage brut/avoir/net séparé
    (OSS_Résumé) sans changer le comportement du XML officiel.

    Args:
        period: période OSS déclarée (ex: "2026-Q1"). Si fournie et reconnue,
            les ventes/avoirs en devise étrangère sont reconvertis en EUR au
            taux BCE du DERNIER JOUR de cette période (Règl. UE 2020/194,
            art. 5 bis) au lieu du taux du jour de la vente déjà figé sur
            `sale.amount_ht` lors de l'import. Si `period` est vide ou non
            reconnu, on retombe sur `sale.amount_ht`/`res.vat_amount` tels
            quels (comportement historique).
    """
    return _aggregate_by_scenario(results, period, scenarios=(Scenario.OSS_B2C,))


def aggregate_ioss_results(results: list[VatResult], period: str = "") -> OssAggType:
    """Agrège les VatResult IOSS_DIRECT — pendant de aggregate_oss_results()
    pour le régime IOSS (import ≤150€, guichet unique séparé, déclaration
    MENSUELLE, art. 369l-x dir. 2006/112/CE). Même structure de retour
    (départ → arrivée → taux) — le "départ" est ici systématiquement un
    pays tiers (hors UE), conservé pour cohérence structurelle même si
    l'IOSS ne segmente pas par pays de départ dans sa déclaration officielle
    (un seul numéro IOSS, ventilation uniquement par pays de consommation
    et taux — voir build_ioss_excel/build_ioss_csv qui aplatissent cette
    dimension à l'export).

    Args:
        period: période IOSS déclarée, typiquement mensuelle (ex: "2026-03"),
            PAS trimestrielle comme pour l'OSS — voir get_oss_rate_date/
            convert_ht_tva_for_oss_period qui savent gérer un format de
            période non reconnu (repli sur sale.amount_ht tel quel).
    """
    return _aggregate_by_scenario(results, period, scenarios=(Scenario.IOSS_DIRECT,))


def _aggregate_by_scenario(
    results: list[VatResult], period: str, scenarios: tuple[Scenario, ...],
) -> OssAggType:
    """Factorisation commune à aggregate_oss_results() et
    aggregate_ioss_results() — même logique d'agrégation départ→arrivée→taux,
    ne diffère que par le filtre de scénario retenu."""
    aggregated: OssAggType = {}

    # Pré-batch des taux BCE nécessaires : sans ça, convert_ht_tva_for_oss_period
    # (via convert_to_currency_for_oss -> get_rate) fait une requête DB
    # individuelle à la PREMIÈRE occurrence de chaque devise/date rencontrée
    # dans la boucle ci-dessous (les occurrences suivantes de la même paire
    # touchent le cache mémoire L1 et sont gratuites — mesuré en prod : 5
    # devises distinctes parmi ~27 lignes OSS_B2C/IOSS_DIRECT ont coûté
    # ~2.9s cumulés en requêtes individuelles séquentielles, alors qu'UNE
    # requête batch groupée suffit). `get_oss_rate_date` est une fonction
    # pure (aucun accès DB) — sûre à appeler ici pour construire l'ensemble
    # des paires à précharger, sans dupliquer la logique de conversion.
    if period:
        _needed_pairs: set[tuple[str, _date]] = set()
        for _res in results:
            if _res.scenario not in scenarios:
                continue
            _src_ccy = _res.sale.original_currency
            if not _src_ccy or _src_ccy == "EUR":
                continue
            try:
                _tx_date = _date.fromisoformat((_res.sale.transaction_date or "")[:10])
            except ValueError:
                _tx_date = _date.today()
            _needed_pairs.add((_src_ccy, get_oss_rate_date(period, _tx_date)))
        if _needed_pairs:
            prefetch_rates(sorted(_needed_pairs))

    for res in results:
        if res.scenario not in scenarios:
            continue

        # fiscal_equivalent_country : Monaco ("MC") n'est pas un État membre UE
        # et n'est jamais un MemberStateOfSupply valide pour le XML officiel
        # OSS — sans cette normalisation, un stock physiquement à Monaco
        # (engine.py::compute_vat traite déjà ce cas en Scenario.OSS_B2C au
        # départ "fiscal" de la France) ferait fuir "MC" tel quel jusque dans
        # le XML déclaratif. Corrigé le 2026-08-26 (angle mort confirmé).
        departure = fiscal_equivalent_country(res.sale.stock_country)   # MemberStateOfSupply (pour OSS) ou Pays tiers (pour IOSS)
        arrival   = res.vat_country          # MemberStateOfConsumption
        rate      = res.vat_rate

        ht, tva = convert_ht_tva_for_oss_period(res, period)

        aggregated.setdefault(departure, {})
        aggregated[departure].setdefault(arrival, {})
        aggregated[departure][arrival].setdefault(
            rate, {
                "ht": Decimal("0.00"), "tva": Decimal("0.00"),
                "ht_vente": Decimal("0.00"), "tva_vente": Decimal("0.00"),
                "ht_remb":  Decimal("0.00"), "tva_remb":  Decimal("0.00"),
                "nb": 0,
            }
        )

        bucket = aggregated[departure][arrival][rate]
        bucket["ht"]  += ht
        bucket["tva"] += tva
        if ht >= 0:
            bucket["ht_vente"]  += ht
            bucket["tva_vente"] += tva
        else:
            bucket["ht_remb"]  += ht
            bucket["tva_remb"] += tva
        bucket["nb"]  += 1

    return aggregated


@dataclass
class OssNegativeBucket:
    """Couple (pays départ → pays destination, taux) dont le solde net OSS
    est négatif sur la période — situation que le portail OSS et le XML
    officiel n'acceptent pas dans le corps principal de la déclaration.

    Survient typiquement quand les avoirs (remboursements) dépassent les
    ventes d'un même pays/taux sur la période — souvent le signe qu'un
    avoir se rapporte en réalité à une vente d'une période antérieure déjà
    déclarée, et qui devrait alors être ventilé dans le bloc
    `CorrectionsOfVatReturns` du XML en référençant la période d'origine
    (Règl. UE 2020/194). L'outil ne peut pas déterminer automatiquement
    cette période d'origine (aucune référence à la vente initiale n'est
    conservée sur l'avoir) — à vérifier et corriger manuellement.
    """
    departure: str
    arrival: str
    vat_rate: Decimal
    base_ht: Decimal
    vat_amount: Decimal


@dataclass
class MatchedRefundCorrection:
    """Un avoir dont l'origine a été rattachée avec CERTITUDE à une vente
    antérieure de la même commande (même sale_id), au sein du même jeu de
    données fourni. Le rattachement se fait UNIQUEMENT sur sale_id identique
    (même couple pays/taux) — jamais par déduction sur order_date, jugé non
    fiable pour générer automatiquement une correction fiscale (voir
    models.py, champ Sale.order_date)."""
    sale_id: str
    origin_period: str          # période OSS d'origine déduite (ex: "2026-Q1")
    base_ht: Decimal
    vat_amount: Decimal
    refund_result: VatResult    # référence à l'objet, pour exclusion précise
                                 # (par identité Python id()) du corps XML
                                 # principal — voir oss_xml.generate_oss_xml.


@dataclass
class NegativeBucketSuggestion:
    """Détail d'un couple (départ, arrivée, taux) en solde négatif, avec la
    part des avoirs qui a pu être rattachée à une vente d'origine identifiée
    (matched, groupée par période d'origine) et la part restée sans
    correspondance (unmatched — à traiter manuellement, comme avant)."""
    bucket: "OssNegativeBucket"
    matched: list[MatchedRefundCorrection]
    unmatched_ht: Decimal
    unmatched_vat_amount: Decimal
    unmatched_count: int

    @property
    def fully_resolved(self) -> bool:
        """True si TOUS les avoirs du couple négatif ont pu être rattachés
        à une origine identifiée — condition nécessaire pour générer
        automatiquement les corrections sans laisser de solde négatif
        résiduel dans le corps principal du XML."""
        return self.unmatched_ht == Decimal("0.00") and self.unmatched_vat_amount == Decimal("0.00")


def _oss_quarter_of(transaction_date: str) -> str:
    """Déduit le trimestre OSS 'YYYY-QN' d'une transaction_date 'YYYY-MM-DD'.
    Retourne '' si la date est vide ou non reconnue."""
    d = (transaction_date or "")[:10]
    if len(d) < 7:
        return ""
    try:
        year = int(d[:4])
        month = int(d[5:7])
    except ValueError:
        return ""
    q = (month - 1) // 3 + 1
    return f"{year}-Q{q}"


def suggest_negative_bucket_corrections(
    results: list[VatResult],
    period: str,
) -> list[NegativeBucketSuggestion]:
    """Pour chaque couple (départ, arrivée, taux) en solde négatif sur la
    période, tente de rattacher chaque avoir constitutif à une vente
    d'origine PRÉSENTE DANS LE MÊME JEU DE DONNÉES `results` (même sale_id,
    même couple pays/taux, montant positif). Ce rattachement n'est possible
    que si le fichier importé couvre aussi la période d'origine de la vente
    créditée — sinon l'avoir reste `unmatched`, exactement comme le
    comportement actuel (blocage manuel).

    N'utilise PAS order_date : seul un sale_id identique, retrouvé dans le
    jeu de données réellement fourni, est considéré comme une preuve
    suffisante pour une correction fiscale automatisée.
    """
    aggregated = aggregate_oss_results(results, period=period)
    negative_buckets = find_oss_negative_buckets(aggregated)
    if not negative_buckets:
        return []

    neg_keys = {(b.departure, b.arrival, b.vat_rate) for b in negative_buckets}

    # Ventes positives disponibles pour matching, indexées par (sale_id, pays, taux)
    positive_by_sale_id: dict[tuple, list[VatResult]] = {}
    refunds_by_bucket: dict[tuple, list[VatResult]] = {}
    for res in results:
        if res.scenario not in (Scenario.OSS_B2C, Scenario.IOSS_DIRECT):
            continue
        # BUGFIX (point #6, README - évolution.md) : `neg_keys` (dérivé de
        # find_oss_negative_buckets(aggregate_oss_results(...))) utilise des
        # pays de départ déjà normalisés via fiscal_equivalent_country() —
        # ex. un stock à Monaco ("MC") y apparaît comme "FR" (convention
        # fiscale franco-monégasque du 18 mai 1963, même normalisation déjà
        # appliquée le 2026-08-26 dans _aggregate_by_scenario). Utiliser ici
        # `res.sale.stock_country` brut ("MC" non normalisé) faisait qu'AUCUNE
        # vente ni avoir à stock Monaco ne matchait jamais `neg_keys`
        # (clé "MC" comparée à une clé "FR") : un compte avec stock Monaco et
        # un couple négatif ne se voyait donc jamais proposer de rattachement
        # automatique, même quand la vente d'origine était présente dans le
        # même fichier — sans erreur visible (le XML restait bloqué par
        # sécurité, mais avec un message "aucun avoir rattaché" trompeur).
        key = (fiscal_equivalent_country(res.sale.stock_country), res.vat_country, res.vat_rate)
        if key not in neg_keys:
            continue
        if res.sale.amount_ht > 0:
            positive_by_sale_id.setdefault(
                (res.sale.sale_id, res.sale.stock_country, res.vat_country, res.vat_rate), []
            ).append(res)
        elif res.sale.amount_ht < 0:
            refunds_by_bucket.setdefault(key, []).append(res)

    suggestions: list[NegativeBucketSuggestion] = []
    for b in negative_buckets:
        key = (b.departure, b.arrival, b.vat_rate)
        matched: list[MatchedRefundCorrection] = []
        unmatched_ht = Decimal("0.00")
        unmatched_vat = Decimal("0.00")
        unmatched_count = 0

        for refund in refunds_by_bucket.get(key, []):
            candidates = positive_by_sale_id.get(
                (refund.sale.sale_id, refund.sale.stock_country, refund.vat_country, refund.vat_rate)
            )
            # BUGFIX (audit du 2026-08-19) : `candidates[0]` prenait la
            # première vente rencontrée dans l'ORDRE D'ITÉRATION de
            # `results` (pas forcément la plus ancienne) quand plusieurs
            # ventes positives partagent le même sale_id (commande
            # multi-articles). On prend désormais explicitement la vente la
            # plus ancienne par transaction_date parmi les candidates, pour
            # que la période d'origine déduite soit déterministe et reflète
            # bien le début de la commande plutôt qu'une ligne arbitraire.
            origin_quarter = (
                _oss_quarter_of(min(candidates, key=lambda c: c.sale.transaction_date or "").sale.transaction_date)
                if candidates else ""
            )
            # On n'accepte le rattachement que si une origine a été trouvée
            # ET qu'elle correspond bien à une période DIFFÉRENTE de la
            # période courante (sinon ce n'est pas un avoir "à cheval",
            # juste un solde négatif normal intra-période — pas notre sujet).
            if origin_quarter and origin_quarter != period:
                matched.append(MatchedRefundCorrection(
                    sale_id=refund.sale.sale_id,
                    origin_period=origin_quarter,
                    base_ht=refund.sale.amount_ht,
                    vat_amount=refund.vat_amount,
                    refund_result=refund,
                ))
            else:
                unmatched_ht += refund.sale.amount_ht
                unmatched_vat += refund.vat_amount
                unmatched_count += 1

        suggestions.append(NegativeBucketSuggestion(
            bucket=b,
            matched=matched,
            unmatched_ht=unmatched_ht,
            unmatched_vat_amount=unmatched_vat,
            unmatched_count=unmatched_count,
        ))

    return suggestions


def find_oss_negative_buckets(aggregated: OssAggType) -> list[OssNegativeBucket]:
    """Liste les couples (départ, arrivée, taux) dont le solde HT ou TVA est négatif."""
    negatives: list[OssNegativeBucket] = []
    for departure, destinations in aggregated.items():
        for arrival, rates in destinations.items():
            for rate, amounts in rates.items():
                if amounts["ht"] < 0 or amounts["tva"] < 0:
                    negatives.append(OssNegativeBucket(
                        departure=departure, arrival=arrival, vat_rate=rate,
                        base_ht=amounts["ht"], vat_amount=amounts["tva"],
                    ))
    return negatives

# Palette couleurs
_BLUE_HEADER = "1F4E79"   # Bleu foncé headers principaux
_BLUE_LIGHT  = "BDD7EE"   # Bleu clair sous-headers
_GREEN_HDR   = "375623"   # Vert foncé onglet B2B
_GREEN_LIGHT = "C6EFCE"   # Vert clair total B2B
_ORANGE_HDR  = "C55A11"   # Orange onglet OSS détail
_ORANGE_LIGHT= "FCE4D6"   # Orange clair
_TOTAL_FILL  = "FFF2CC"   # Jaune ligne totaux
_WHITE       = "FFFFFF"
_GREY_ROW    = "F2F2F2"

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _country_name(code: str) -> str:
    """Nom localisé du pays (clé i18n `country_XX`).
    Remplace l'ancien dict COUNTRY_NAMES codé en dur en français."""
    return country_label(code)


@dataclass
class OssCountryLine:
    country: str
    country_name: str
    vat_rate: Decimal
    base_ht: Decimal          # NET (vente + avoir) — clé historique
    vat_amount: Decimal       # NET (vente + avoir) — clé historique
    nb_transactions: int
    base_ht_vente: Decimal = _ZERO   # brut, ventes seules
    vat_vente: Decimal      = _ZERO   # brut, ventes seules
    base_ht_remb: Decimal   = _ZERO   # avoirs seuls (négatif)
    vat_remb: Decimal       = _ZERO   # avoirs seuls (négatif)


@dataclass
class B2bLine:
    sale_id: str
    buyer_vat_number: str
    buyer_country: str
    country_name: str
    amount_ht: Decimal
    transaction_date: str


@dataclass
class OssExportData:
    oss_by_country: List[OssCountryLine]
    oss_details: List[tuple]   # (VatResult, ht_converti, tva_converti) — voir convert_ht_tva_for_oss_period
    b2b_lines: List[B2bLine]
    period: str
    total_oss_ht: Decimal = _ZERO
    total_oss_vat: Decimal = _ZERO
    total_b2b_ht: Decimal = _ZERO


def _aggregate(results: List[VatResult], period: str = "") -> OssExportData:
    """Agrège les VatResult en données prêtes pour l'export.

    S'appuie sur aggregate_oss_results() — source unique de vérité partagée
    avec oss_xml.py pour garantir la cohérence entre l'Excel URSSAF et le XML OSS.
    """
    oss_agg = aggregate_oss_results(results, period=period)
    b2b_results = [r for r in results if r.scenario == Scenario.B2B_REVERSE_CHARGE]

    # Aplatissement de la structure hiérarchique départ→arrivée→taux
    # en une liste plate par (pays_destination, taux) pour l'Excel URSSAF.
    # Note : l'Excel OSS consolide TOUS les pays de départ — la DGFiP attend
    # une vue par pays de consommation (pas de départ) dans l'état OSS FR.
    country_map: dict[tuple[str, Decimal], dict] = {}

    for departure, destinations in oss_agg.items():
        for arrival, rates in destinations.items():
            for rate, amounts in rates.items():
                key = (arrival, rate)
                if key not in country_map:
                    country_map[key] = {
                        "country": arrival,
                        "country_name": _country_name(arrival),
                        "vat_rate": rate,
                        "base_ht": _ZERO,
                        "vat_amount": _ZERO,
                        "base_ht_vente": _ZERO,
                        "vat_vente": _ZERO,
                        "base_ht_remb": _ZERO,
                        "vat_remb": _ZERO,
                        "nb": 0,
                    }
                country_map[key]["base_ht"]       += amounts["ht"]
                country_map[key]["vat_amount"]     += amounts["tva"]
                country_map[key]["base_ht_vente"]  += amounts["ht_vente"]
                country_map[key]["vat_vente"]      += amounts["tva_vente"]
                country_map[key]["base_ht_remb"]   += amounts["ht_remb"]
                country_map[key]["vat_remb"]       += amounts["tva_remb"]
                country_map[key]["nb"]             += amounts["nb"]

    # Reconstruire la liste de détail OSS depuis les résultats d'origine, en
    # appliquant la MÊME reconversion BCE de clôture de période que le Résumé
    # (convert_ht_tva_for_oss_period) — auparavant le détail affichait le
    # montant au taux du jour de la vente, différent du Résumé pour toute
    # vente en devise étrangère.
    oss_detail_results = [
        (r, *convert_ht_tva_for_oss_period(r, period))
        for r in results if r.scenario == Scenario.OSS_B2C
    ]

    oss_lines = [
        OssCountryLine(
            country=v["country"],
            country_name=v["country_name"],
            vat_rate=v["vat_rate"],
            base_ht=v["base_ht"],
            vat_amount=v["vat_amount"],
            nb_transactions=v["nb"],
            base_ht_vente=v["base_ht_vente"],
            vat_vente=v["vat_vente"],
            base_ht_remb=v["base_ht_remb"],
            vat_remb=v["vat_remb"],
        )
        for v in sorted(country_map.values(), key=lambda x: x["country"])
    ]

    b2b_lines = [
        B2bLine(
            # sale_id/buyer_vat_number ne servent ici qu'à l'affichage (Excel +
            # CSV plus bas) — jamais à un matching -> _safe() sans risque de
            # casser une comparaison ailleurs (vérifié : seul res.sale.sale_id
            # brut, hors B2bLine, est utilisé pour le matching remboursements).
            sale_id=_safe(getattr(r.sale, "display_id", "") or r.sale.sale_id),
            buyer_vat_number=_safe(r.sale.buyer_vat_number),
            buyer_country=r.sale.buyer_country,
            country_name=_country_name(r.sale.buyer_country),
            amount_ht=r.sale.amount_ht,
            transaction_date=r.sale.transaction_date,
        )
        for r in b2b_results
    ]

    return OssExportData(
        oss_by_country=oss_lines,
        oss_details=oss_detail_results,
        b2b_lines=b2b_lines,
        period="",
        total_oss_ht=sum((l.base_ht for l in oss_lines), _ZERO),
        total_oss_vat=sum((l.vat_amount for l in oss_lines), _ZERO),
        total_b2b_ht=sum((l.amount_ht for l in b2b_lines), _ZERO),
    )


def _merge(ws, cell_range: str) -> None:
    """Fusionne des cellules en mode `write_only` (l'API `ws.merge_cells()`
    normale n'existe pas sur `WriteOnlyWorksheet` — vérifié empiriquement
    compatible en passant directement par `ws.merged_cells.ranges`, qui
    est la structure lue par openpyxl à la sauvegarde quel que soit le mode)."""
    ws.merged_cells.ranges.add(CellRange(cell_range))


_FORMULA_LEADING_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _safe(value):
    """Neutralise une injection de formule Excel/CSV (OWASP CSV Injection)
    sur une valeur texte issue du fichier Amazon importé (sale_id,
    display_id, numéro de TVA acheteur...) — non fiable par nature.
    Même logique que `excel_report.py::_safe`, dupliquée ici car ce module
    n'importe pas excel_report (pas de dépendance croisée souhaitée).
    Ne jamais appliquer à des formules internes construites par l'appli."""
    if isinstance(value, str) and value and value[0] in _FORMULA_LEADING_CHARS:
        return "'" + value
    return value


def _wcell(ws, value, font=None, fill=None, alignment=None, number_format=None, border=None):
    """Construit une cellule stylée prête à être placée dans une ligne et
    ajoutée via `ws.append([...])` — seule API d'écriture disponible en
    mode `write_only` (pas d'accès aléatoire type `ws.cell(row=, column=)`).
    `WriteOnlyCell` fonctionne aussi bien en `Workbook()` normal qu'en
    `Workbook(write_only=True)` (même pattern que `excel_report.py::_wcell`)."""
    cell = WriteOnlyCell(ws, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format
    if border is not None:
        cell.border = border
    return cell


def _hdr_cell(ws, value: str, bg: str, fg: str = _WHITE, bold: bool = True, size: int = 10):
    return _wcell(
        ws, value,
        font=Font(bold=bold, color=fg, name="Arial", size=size),
        fill=PatternFill("solid", start_color=bg),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=_BORDER,
    )


def _data_cell(ws, value, fmt: str = None, zebra: bool = False, alignment: Alignment = None):
    return _wcell(
        ws, value,
        font=Font(name="Arial", size=9),
        fill=PatternFill("solid", start_color=_GREY_ROW) if zebra else None,
        alignment=alignment or Alignment(vertical="center"),
        number_format=fmt,
        border=_BORDER,
    )


def _total_cell(ws, value, fmt: str = None, alignment: Alignment = None):
    return _wcell(
        ws, value,
        font=Font(bold=True, name="Arial", size=9),
        fill=PatternFill("solid", start_color=_TOTAL_FILL),
        alignment=alignment or Alignment(vertical="center"),
        number_format=fmt,
        border=_BORDER,
    )


def _build_oss_resume(
        wb: Workbook, data: "OssExportData | IossExportData", period: str,
        sheet_name: str = "OSS_Résumé",
        title_key: str = "oss_export_title", subtitle_key: str = "oss_export_subtitle",
        total_label_key: str = "oss_total_countries", footer_key: str = "oss_footer_note",
        lines: List["OssCountryLine"] | None = None,
):
    """Construit l'onglet Résumé pays/taux — factorisé pour être réutilisé
    tel quel par _build_ioss_resume() (même structure, juste les clés i18n/
    nom d'onglet qui changent). `lines` permet de passer une autre liste
    que `data.oss_by_country` (ex: `data.ioss_by_country`) sans dupliquer
    tout le corps de la fonction."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    _lines = lines if lines is not None else data.oss_by_country

    # Largeurs de colonnes : DOIVENT être fixées avant le tout premier
    # `append` (vérifié empiriquement en write_only — contrairement au mode
    # normal, ce n'est pas juste "avant le premier append de LA colonne").
    widths = [12, 22, 10, 16, 15, 16, 15, 16, 15, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Titre
    ws.row_dimensions[1].height = 28
    t = _wcell(
        ws, _(title_key, period=period),
        font=Font(bold=True, size=13, color=_WHITE, name="Arial"),
        fill=PatternFill("solid", start_color=_BLUE_HEADER),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append([t])
    _merge(ws, "A1:J1")

    sub = _wcell(
        ws, _(subtitle_key),
        font=Font(italic=True, size=9, color="595959", name="Arial"),
        alignment=Alignment(horizontal="center"),
    )
    ws.append([sub])
    _merge(ws, "A2:J2")

    # Headers colonnes
    headers = [
        _("oss_col_country_code"), _("oss_col_country"), _("oss_col_vat_rate"),
        _("oss_col_base_sales"), _("oss_col_vat_sales"),
        _("oss_col_base_refunds"), _("oss_col_vat_refunds"),
        _("oss_col_base_net"), _("oss_col_vat_net"),
        _("oss_col_nb_tx")
    ]
    ws.row_dimensions[3].height = 20
    ws.append([_hdr_cell(ws, h, _BLUE_LIGHT, fg="1F4E79", size=9) for h in headers])

    # Données
    for i, line in enumerate(_lines):
        r = i + 4
        zebra = i % 2 == 1
        ws.row_dimensions[r].height = 16
        ws.append([
            _data_cell(ws, line.country, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, line.country_name, zebra=zebra),
            _data_cell(ws, float(line.vat_rate) / 100, fmt="0.0%", zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, float(line.base_ht_vente), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(line.vat_vente), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(line.base_ht_remb), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(line.vat_remb), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(line.base_ht), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(line.vat_amount), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, line.nb_transactions, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
        ])

    # Ligne total
    n = len(_lines)
    total_row = n + 4
    ws.row_dimensions[total_row].height = 18
    ws.append([
        _total_cell(ws, _("TOTAL"), alignment=Alignment(horizontal="center", vertical="center")),
        _total_cell(ws, _(total_label_key, count=n)),
        _total_cell(ws, ""),
        _total_cell(ws, f"=SUM(D4:D{total_row-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, f"=SUM(E4:E{total_row-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, f"=SUM(F4:F{total_row-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, f"=SUM(G4:G{total_row-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, f"=SUM(H4:H{total_row-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, f"=SUM(I4:I{total_row-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, f"=SUM(J4:J{total_row-1})", alignment=Alignment(horizontal="center", vertical="center")),
    ])

    # Note de bas de page
    note_row = total_row + 2
    n_cell = _wcell(ws, _(footer_key), font=Font(italic=True, size=8, color="C00000", name="Arial"))
    # Ligne intermédiaire vide (note_row = total_row + 2 dans l'original,
    # donc une ligne blanche entre le total et la note).
    ws.append([])
    ws.append([n_cell])
    _merge(ws, f"A{note_row}:J{note_row}")


def _build_oss_detail(wb: Workbook, data: OssExportData):
    ws = wb.create_sheet("OSS_Détail")
    ws.sheet_view.showGridLines = False

    widths = [20, 14, 10, 14, 16, 16, 11, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25
    t = _wcell(
        ws, _("oss_detail_title"),
        font=Font(bold=True, size=12, color=_WHITE, name="Arial"),
        fill=PatternFill("solid", start_color=_ORANGE_HDR),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append([t])
    _merge(ws, "A1:H1")

    headers = [
        _("oss_detail_col_id"), _("oss_detail_col_date"), _("oss_detail_col_stock"),
        _("oss_detail_col_dest"), _("oss_detail_col_dest_name"),
        _("oss_detail_col_base_ht"), _("oss_detail_col_vat_rate"), _("oss_detail_col_vat_amount")
    ]
    ws.row_dimensions[2].height = 18
    ws.append([_hdr_cell(ws, h, _ORANGE_LIGHT, fg=_ORANGE_HDR, size=9) for h in headers])

    # data.oss_details contient des tuples (VatResult, ht_converti, tva_converti) —
    # ht/tva déjà reconvertis au taux BCE de clôture de période (même valeur
    # que celle agrégée dans OSS_Résumé, voir convert_ht_tva_for_oss_period).
    for i, (r, ht, tva) in enumerate(data.oss_details):
        row = i + 3
        zebra = i % 2 == 1
        ws.row_dimensions[row].height = 15
        ws.append([
            _data_cell(ws, _safe(getattr(r.sale, "display_id", "") or r.sale.sale_id), zebra=zebra),
            _data_cell(ws, r.sale.transaction_date, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, r.sale.stock_country, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, r.sale.buyer_country, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, _country_name(r.sale.buyer_country), zebra=zebra),
            _data_cell(ws, float(ht), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(r.vat_rate) / 100, fmt="0.0%", zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, float(tva), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
        ])

    # Totaux
    n = len(data.oss_details)
    tr = n + 3
    ws.append([
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, _("TOTAL")),
        _total_cell(ws, f"=SUM(F3:F{tr-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, ""),
        _total_cell(ws, f"=SUM(H3:H{tr-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
    ])


def _b2b_month_key(transaction_date: str) -> str:
    """Clé de regroupement mensuel (YYYY-MM) à partir d'une date normalisée
    YYYY-MM-DD (voir parsers/amazon/detect.py::parse_date). Chaîne vide
    (date manquante/invalide) regroupée à part, triée en dernier."""
    return transaction_date[:7] if len(transaction_date) >= 7 else ""


def _build_b2b_recap(wb: Workbook, data: OssExportData, period: str):
    """Onglet B2B_Recap avec sous-totaux mensuels : l'état récapitulatif
    des clients (DES, art. 289 B CGI) est une obligation MENSUELLE, alors
    que `period` (l'export choisi par l'utilisateur, ex. un trimestre) peut
    couvrir plusieurs mois. On garde un seul onglet — trié chronologiquement,
    avec un bandeau + un sous-total par mois — plutôt que fragmenter en
    plusieurs fichiers, pour rester lisible en un coup d'œil tout en donnant
    les montants exacts à reporter mois par mois."""
    ws = wb.create_sheet("B2B_Recap")
    ws.sheet_view.showGridLines = False

    widths = [20, 14, 22, 12, 20, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25
    t = _wcell(
        ws, _("b2b_recap_title", period=period),
        font=Font(bold=True, size=12, color=_WHITE, name="Arial"),
        fill=PatternFill("solid", start_color=_GREEN_HDR),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append([t])
    _merge(ws, "A1:F1")

    sub = _wcell(
        ws, _("b2b_recap_subtitle"),
        font=Font(italic=True, size=9, color="595959", name="Arial"),
        alignment=Alignment(horizontal="center"),
    )
    ws.append([sub])
    _merge(ws, "A2:F2")

    headers = [
        _("b2b_col_id"), _("b2b_col_date"), _("b2b_col_vat_number"),
        _("b2b_col_country_code"), _("b2b_col_buyer_country"), _("b2b_col_amount_ht")
    ]
    ws.row_dimensions[3].height = 18
    ws.append([_hdr_cell(ws, h, _GREEN_LIGHT, fg=_GREEN_HDR, size=9) for h in headers])

    # Tri chronologique puis regroupement par mois (YYYY-MM) ; les lignes
    # sans date exploitable sont regroupées à part, en dernier.
    _sorted_lines = sorted(data.b2b_lines, key=lambda l: (_b2b_month_key(l.transaction_date) == "", _b2b_month_key(l.transaction_date), l.transaction_date))

    _months: dict[str, list] = {}
    for line in _sorted_lines:
        _months.setdefault(_b2b_month_key(line.transaction_date), []).append(line)

    row = 3  # dernière ligne écrite = en-têtes (row 3)
    zebra_idx = 0
    for month_key, lines in _months.items():
        month_label = month_key if month_key else _("b2b_unknown_date")

        row += 1
        ws.row_dimensions[row].height = 16
        ws.append([_hdr_cell(ws, _("b2b_month_group_header", month=month_label), "E2EFDA", fg="375623", size=8)])
        _merge(ws, f"A{row}:F{row}")

        month_total = _ZERO
        for line in lines:
            row += 1
            zebra = zebra_idx % 2 == 1
            zebra_idx += 1
            ws.row_dimensions[row].height = 15
            ws.append([
                _data_cell(ws, line.sale_id, zebra=zebra),
                _data_cell(ws, line.transaction_date, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
                _data_cell(ws, line.buyer_vat_number or "—", zebra=zebra),
                _data_cell(ws, line.buyer_country, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
                _data_cell(ws, line.country_name, zebra=zebra),
                _data_cell(ws, float(line.amount_ht), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            ])
            month_total += line.amount_ht

        row += 1
        ws.row_dimensions[row].height = 16
        ws.append([
            _total_cell(ws, ""),
            _total_cell(ws, ""),
            _total_cell(ws, ""),
            _total_cell(ws, ""),
            _total_cell(ws, _("b2b_month_subtotal", month=month_label)),
            _total_cell(ws, float(month_total), fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        ])

    tr = row + 1
    ws.append([
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, _("TOTAL HT")),
        _total_cell(ws, float(data.total_b2b_ht), fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
    ])

    note_row = tr + 2
    n_cell = _wcell(ws, _("b2b_footer_note"), font=Font(italic=True, size=8, color="C00000", name="Arial"))
    ws.append([])
    ws.append([n_cell])
    _merge(ws, f"A{note_row}:F{note_row}")


@dataclass
class IossExportData:
    """Pendant de OssExportData pour le régime IOSS (import ≤150€, guichet
    séparé, déclaration MENSUELLE — voir aggregate_ioss_results)."""
    ioss_by_country: List[OssCountryLine]
    ioss_details: List[tuple]   # (VatResult, ht_converti, tva_converti)
    period: str
    total_ioss_ht: Decimal = _ZERO
    total_ioss_vat: Decimal = _ZERO


def _aggregate_ioss(results: List[VatResult], period: str = "") -> IossExportData:
    """Pendant de _aggregate() pour l'IOSS — s'appuie sur
    aggregate_ioss_results() (Scenario.IOSS_DIRECT uniquement, jamais
    OSS_B2C). period attendu au format mensuel (ex: "2026-03"), pas
    trimestriel."""
    ioss_agg = aggregate_ioss_results(results, period=period)

    country_map: dict[tuple[str, Decimal], dict] = {}
    for departure, destinations in ioss_agg.items():
        for arrival, rates in destinations.items():
            for rate, amounts in rates.items():
                key = (arrival, rate)
                if key not in country_map:
                    country_map[key] = {
                        "country": arrival,
                        "country_name": _country_name(arrival),
                        "vat_rate": rate,
                        "base_ht": _ZERO,
                        "vat_amount": _ZERO,
                        "base_ht_vente": _ZERO,
                        "vat_vente": _ZERO,
                        "base_ht_remb": _ZERO,
                        "vat_remb": _ZERO,
                        "nb": 0,
                    }
                country_map[key]["base_ht"]       += amounts["ht"]
                country_map[key]["vat_amount"]     += amounts["tva"]
                country_map[key]["base_ht_vente"]  += amounts["ht_vente"]
                country_map[key]["vat_vente"]      += amounts["tva_vente"]
                country_map[key]["base_ht_remb"]   += amounts["ht_remb"]
                country_map[key]["vat_remb"]       += amounts["tva_remb"]
                country_map[key]["nb"]             += amounts["nb"]

    ioss_detail_results = [
        (r, *convert_ht_tva_for_oss_period(r, period))
        for r in results if r.scenario == Scenario.IOSS_DIRECT
    ]

    ioss_lines = [
        OssCountryLine(
            country=v["country"],
            country_name=v["country_name"],
            vat_rate=v["vat_rate"],
            base_ht=v["base_ht"],
            vat_amount=v["vat_amount"],
            nb_transactions=v["nb"],
            base_ht_vente=v["base_ht_vente"],
            vat_vente=v["vat_vente"],
            base_ht_remb=v["base_ht_remb"],
            vat_remb=v["vat_remb"],
        )
        for v in sorted(country_map.values(), key=lambda x: x["country"])
    ]

    return IossExportData(
        ioss_by_country=ioss_lines,
        ioss_details=ioss_detail_results,
        period="",
        total_ioss_ht=sum((l.base_ht for l in ioss_lines), _ZERO),
        total_ioss_vat=sum((l.vat_amount for l in ioss_lines), _ZERO),
    )


def _build_ioss_resume(wb: Workbook, data: IossExportData, period: str):
    _build_oss_resume(
        wb, data, period,
        sheet_name="IOSS_Résumé",
        title_key="ioss_export_title", subtitle_key="ioss_export_subtitle",
        total_label_key="oss_total_countries", footer_key="ioss_footer_note",
        lines=data.ioss_by_country,
    )


def _build_ioss_detail(wb: Workbook, data: IossExportData):
    ws = wb.create_sheet("IOSS_Détail")
    ws.sheet_view.showGridLines = False

    widths = [20, 14, 10, 14, 16, 16, 11, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25
    t = _wcell(
        ws, _("ioss_detail_title"),
        font=Font(bold=True, size=12, color=_WHITE, name="Arial"),
        fill=PatternFill("solid", start_color=_ORANGE_HDR),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append([t])
    _merge(ws, "A1:H1")

    headers = [
        _("oss_detail_col_id"), _("oss_detail_col_date"), _("oss_detail_col_stock"),
        _("oss_detail_col_dest"), _("oss_detail_col_dest_name"),
        _("oss_detail_col_base_ht"), _("oss_detail_col_vat_rate"), _("oss_detail_col_vat_amount")
    ]
    ws.row_dimensions[2].height = 18
    ws.append([_hdr_cell(ws, h, _ORANGE_LIGHT, fg=_ORANGE_HDR, size=9) for h in headers])

    for i, (r, ht, tva) in enumerate(data.ioss_details):
        row = i + 3
        zebra = i % 2 == 1
        ws.row_dimensions[row].height = 15
        ws.append([
            _data_cell(ws, _safe(getattr(r.sale, "display_id", "") or r.sale.sale_id), zebra=zebra),
            _data_cell(ws, r.sale.transaction_date, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, r.sale.stock_country, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, r.sale.buyer_country, zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, _country_name(r.sale.buyer_country), zebra=zebra),
            _data_cell(ws, float(ht), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
            _data_cell(ws, float(r.vat_rate) / 100, fmt="0.0%", zebra=zebra, alignment=Alignment(horizontal="center", vertical="center")),
            _data_cell(ws, float(tva), fmt='#,##0.00 "€"', zebra=zebra, alignment=Alignment(horizontal="right", vertical="center")),
        ])

    n = len(data.ioss_details)
    tr = n + 3
    ws.append([
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, ""),
        _total_cell(ws, _("TOTAL")),
        _total_cell(ws, f"=SUM(F3:F{tr-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
        _total_cell(ws, ""),
        _total_cell(ws, f"=SUM(H3:H{tr-1})", fmt='#,##0.00 "€"', alignment=Alignment(horizontal="right", vertical="center")),
    ])


def build_ioss_excel(results: List[VatResult], output_path: str, period: str = "") -> None:
    """Export IOSS dédié (mensuel) — pendant de build_oss_excel() pour le
    régime IOSS. ⚠️ Ce n'est PAS un XML officiel homologué (contrairement à
    generate_oss_xml pour l'OSS union scheme) : le format XML IOSS (import
    scheme, Règl. UE 2020/194 art. 8 bis et s.) n'est pas implémenté ici,
    faute de spécification technique disponible au moment de ce correctif.
    Cet export permet au moins de ne plus PERDRE les montants IOSS collectés
    (silencieusement exclus du XML OSS avant ce correctif) et de préparer la
    déclaration IOSS mensuelle manuellement en attendant une implémentation
    XML dédiée."""
    data = _aggregate_ioss(results, period=period)
    wb = Workbook(write_only=True)
    _build_ioss_resume(wb, data, period)
    _build_ioss_detail(wb, data)
    wb.save(output_path)


def build_ioss_csv(results: List[VatResult], period: str = "") -> str:
    """Export CSV IOSS (mensuel), format FR (virgule décimale, ';' séparateur)
    — même convention que build_oss_csv()."""
    data = _aggregate_ioss(results, period=period)
    lines = ["Pays;Nom pays;Taux TVA;Base HT ventes;TVA ventes;Base HT avoirs;TVA avoirs;Base HT nette;TVA nette;Nb transactions"]
    for line in data.ioss_by_country:
        lines.append(";".join([
            line.country, line.country_name, _fmt_dec(line.vat_rate),
            _fmt_dec(line.base_ht_vente), _fmt_dec(line.vat_vente),
            _fmt_dec(line.base_ht_remb), _fmt_dec(line.vat_remb),
            _fmt_dec(line.base_ht), _fmt_dec(line.vat_amount),
            str(line.nb_transactions),
        ]))
    return "\n".join(lines)


def _fmt_dec(value: Optional[Decimal]) -> str:
    """Formate un Decimal pour le CSV FR (virgule décimale), en tolérant
    None/valeurs déjà quantizées sans planter sur un `.replace()` appelé sur
    autre chose qu'une string bien formée."""
    if value is None:
        value = _ZERO
    return f"{value:.2f}".replace(".", ",")


def build_oss_excel(
    results: List[VatResult],
    output_path: str | Path,
    period: str = "",
    data: "OssExportData | None" = None,
) -> Path:
    """Génère le fichier Excel multi-onglets OSS uniquement.

    Args:
        results: liste de VatResult issus du moteur.
        output_path: chemin de sortie du fichier .xlsx.
        period: libellé de la période (ex: "2024-T1", "Mars 2024").
        data: OssExportData déjà agrégé (évite un recalcul si l'appelant l'a
              déjà — voir build_oss_export). Si omis, agrège `results` ici.

    Returns:
        Path du fichier généré.
    """
    if data is None:
        data = _aggregate(results, period=period)
    data.period = period

    wb = Workbook(write_only=True)
    # Contrairement au mode normal, `Workbook(write_only=True)` ne crée pas
    # de feuille par défaut : pas de `wb.remove(wb.active)` nécessaire.

    _build_oss_resume(wb, data, period)
    _build_oss_detail(wb, data)

    output_path = Path(output_path)
    wb.save(str(output_path))
    return output_path


def build_b2b_excel(
    results: List[VatResult],
    output_path: str | Path,
    period: str = "",
    data: "OssExportData | None" = None,
) -> Path:
    """Génère le fichier Excel pour les livraisons B2B (État récapitulatif)."""
    if data is None:
        data = _aggregate(results, period=period)

    wb = Workbook(write_only=True)
    _build_b2b_recap(wb, data, period)

    output_path = Path(output_path)
    wb.save(str(output_path))
    return output_path


def build_oss_csv(
    results: List[VatResult],
    period: str = "",
    data: "OssExportData | None" = None,
) -> tuple[bytes, bytes]:
    """Génère les deux CSV : OSS URSSAF et B2B récapitulatif.

    Args:
        data: OssExportData déjà agrégé (voir build_oss_export) ; si omis,
              agrège `results` ici (utilisable en standalone).
    """
    if data is None:
        data = _aggregate(results, period=period)

    # --- CSV OSS URSSAF ---
    oss_buf = io.StringIO()
    oss_writer = csv.writer(oss_buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    oss_writer.writerow([_("oss_csv_title", period=period)])
    oss_writer.writerow([])
    oss_writer.writerow([_("oss_col_country_code"), _("oss_col_country"), _("oss_csv_col_rate_pct"), _("oss_csv_col_base_eur"), _("oss_csv_col_vat_eur"), _("oss_col_nb_tx")])
    for line in data.oss_by_country:
        oss_writer.writerow([
            line.country,
            line.country_name,
            _fmt_dec(line.vat_rate),
            _fmt_dec(line.base_ht),
            _fmt_dec(line.vat_amount),
            line.nb_transactions,
        ])
    oss_writer.writerow([])
    oss_writer.writerow([
        _("TOTAL"), "",
        "",
        _fmt_dec(data.total_oss_ht),
        _fmt_dec(data.total_oss_vat),
        sum(l.nb_transactions for l in data.oss_by_country),
    ])

    # --- CSV B2B ---
    b2b_buf = io.StringIO()
    b2b_writer = csv.writer(b2b_buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    b2b_writer.writerow([_("b2b_csv_title", period=period)])
    b2b_writer.writerow([])
    b2b_writer.writerow([_("b2b_col_id"), _("b2b_col_date"), _("b2b_col_vat_number"), _("b2b_col_country_code"), _("b2b_col_buyer_country"), _("b2b_csv_col_amount_eur")])

    # Meme regroupement mensuel que l'onglet Excel B2B_Recap (obligation DES
    # mensuelle, art. 289 B CGI) — voir _build_b2b_recap pour le detail.
    _sorted_b2b = sorted(data.b2b_lines, key=lambda l: (_b2b_month_key(l.transaction_date) == "", _b2b_month_key(l.transaction_date), l.transaction_date))
    _b2b_months: dict[str, list] = {}
    for line in _sorted_b2b:
        _b2b_months.setdefault(_b2b_month_key(line.transaction_date), []).append(line)

    for month_key, lines in _b2b_months.items():
        month_label = month_key if month_key else _("b2b_unknown_date")
        b2b_writer.writerow([_("b2b_month_group_header", month=month_label)])
        month_total = _ZERO
        for line in lines:
            b2b_writer.writerow([
                line.sale_id,
                line.transaction_date,
                line.buyer_vat_number or "",
                line.buyer_country,
                line.country_name,
                _fmt_dec(line.amount_ht),
            ])
            month_total += line.amount_ht
        b2b_writer.writerow(["", "", "", "", _("b2b_month_subtotal", month=month_label), _fmt_dec(month_total)])

    b2b_writer.writerow([])
    b2b_writer.writerow([_("TOTAL"), "", "", "", "", _fmt_dec(data.total_b2b_ht)])

    # UTF-8 BOM pour compatibilité Excel
    oss_bytes = ("\ufeff" + oss_buf.getvalue()).encode("utf-8")
    b2b_bytes = ("\ufeff" + b2b_buf.getvalue()).encode("utf-8")
    return oss_bytes, b2b_bytes


def build_oss_export(
    results: List[VatResult],
    output_dir: str | Path,
    period: str = "",
) -> tuple[Path, bytes, bytes]:
    """Point d'entrée principal : génère Excel + les deux CSV.

    N'agrège `results` qu'une seule fois (au lieu de deux : une fois pour
    l'Excel, une fois pour le CSV) et réutilise le même OssExportData pour
    les deux exports.

    Returns:
        Tuple (xlsx_path, oss_csv_bytes, b2b_csv_bytes).
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    data = _aggregate(results, period=period)
    xlsx_path = build_oss_excel(results, output_dir / "etat_recapitulatif_oss.xlsx", period, data=data)
    oss_csv, b2b_csv = build_oss_csv(results, period, data=data)
    return xlsx_path, oss_csv, b2b_csv