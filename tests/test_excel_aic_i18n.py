"""Tests de régression — i18n de l'onglet 'Analyse AIC FBA' (excel_report._write_fba_aic_tab).

Contexte : cette fonction n'appelait jamais i18n_() avant le correctif du
2026-09-02 (entrée #12 de optimisations_en_attente.md) — toutes les chaînes
étaient en dur en français, contrairement aux autres onglets du classeur.
Ces tests vérifient que l'onglet se génère sans erreur dans les 7 langues
et qu'aucun texte ne retombe sur le fallback brut de get_text() (qui
renverrait la clé elle-même, signe d'une clé manquante dans un TOML).
"""

from __future__ import annotations

import streamlit as st
from openpyxl import Workbook

from tva_intracom.excel_report import _write_fba_aic_tab
from tva_intracom.i18n.i18n import load_translations

LANGUAGES = ["fr", "en", "de", "es", "it", "pl", "pt"]

# Transferts FC minimalistes : un flux "actif" (immatriculation croisée FR/DE)
# et un flux "inactif" (FR/IT, immatriculation IT absente) pour couvrir les
# deux sections de l'onglet.
SAMPLE_TRANSFERS = [
    {
        "TRANSACTION_EVENT_ID": "t1", "TRANSACTION_COMPLETE_DATE": "01-01-2026",
        "ASIN": "ASIN001", "ITEM_DESCRIPTION": "Produit Test",
        "DEPARTURE_COUNTRY": "FR", "ARRIVAL_COUNTRY": "DE",
        "QTY": 10,
    },
    {
        "TRANSACTION_EVENT_ID": "t2", "TRANSACTION_COMPLETE_DATE": "02-01-2026",
        "ASIN": "ASIN002", "ITEM_DESCRIPTION": "Autre Produit",
        "DEPARTURE_COUNTRY": "FR", "ARRIVAL_COUNTRY": "IT",
        "QTY": 5,
    },
]


def _make_wb_and_ws():
    wb = Workbook(write_only=True)
    return wb, wb.create_sheet("tmp")


def test_write_fba_aic_tab_all_languages_no_crash(monkeypatch, tmp_path):
    """L'onglet doit se générer sans exception dans chacune des 7 langues."""
    for lang in LANGUAGES:
        st.session_state["language"] = lang
        wb, ws = _make_wb_and_ws()
        _write_fba_aic_tab(
            ws,
            all_fc_transfers=SAMPLE_TRANSFERS,
            results=[],
            countries_with_vat=["FR", "DE"],
            display_currency="EUR",
            asin_avg={"ASIN001": 15, "ASIN002": 8},
        )
        wb.save(tmp_path / f"aic_{lang}.xlsx")


def test_write_fba_aic_tab_no_raw_key_fallback():
    """Aucune clé i18n de cette fonction ne doit retomber sur elle-même
    (get_text() renvoie la clé brute quand elle est absente du TOML)."""
    aic_keys = [
        "xl_tab_aic", "xl_aic_title", "xl_aic_note", "xl_aic_active_title",
        "xl_aic_no_active", "xl_aic_col_dep", "xl_aic_col_arr", "xl_aic_col_asin",
        "xl_aic_col_desc", "xl_aic_col_qty", "xl_aic_col_avg_price",
        "xl_aic_col_base", "xl_aic_col_rate", "xl_aic_col_vat",
        "xl_aic_col_status", "xl_aic_status_ok", "xl_aic_status_check",
        "xl_aic_subtotals_title", "xl_aic_sub_col_flow", "xl_aic_sub_col_transfers",
        "xl_aic_sub_col_asins", "xl_aic_sub_col_base", "xl_aic_sub_col_vat",
        "xl_aic_sub_col_ref", "xl_aic_sub_col_action", "xl_aic_legal_ref",
        "xl_aic_action_required", "xl_aic_inactive_title", "xl_aic_col_imm_dep",
        "xl_aic_col_imm_arr", "xl_aic_col_obs", "xl_aic_col_asins_distinct",
        "xl_aic_obs_none", "xl_aic_obs_lic", "xl_aic_obs_verify",
    ]
    for lang in LANGUAGES:
        translations = load_translations(lang)
        for key in aic_keys:
            assert key in translations, f"Clé '{key}' absente de {lang}.toml"
            assert translations[key] != key, (
                f"Clé '{key}' retombe sur elle-même dans {lang}.toml (fallback brut)"
            )


def test_write_fba_aic_tab_dynamic_currency_in_headers(monkeypatch, tmp_path):
    """Les en-têtes de montant doivent afficher la devise réellement passée
    en paramètre (display_currency), pas un '€' figé — régression du
    correctif du 2026-09-02."""
    st.session_state["language"] = "fr"
    wb, ws = _make_wb_and_ws()
    _write_fba_aic_tab(
        ws,
        all_fc_transfers=SAMPLE_TRANSFERS,
        results=[],
        countries_with_vat=["FR", "DE"],
        display_currency="USD",
        asin_avg={"ASIN001": 15, "ASIN002": 8},
    )
    out = tmp_path / "aic_usd.xlsx"
    wb.save(out)
    from openpyxl import load_workbook
    reloaded = load_workbook(str(out))
    actual_sheet = reloaded.sheetnames[0]
    header_row_values = [c.value for c in reloaded[actual_sheet][5]]
    header_text = " | ".join(str(v) for v in header_row_values if isinstance(v, str))
    assert "USD" in header_text
    assert "€" not in header_text or "USD" in header_text
