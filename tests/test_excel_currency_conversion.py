"""Test de non-régression pour l'optimisation de _write_details_tab
(voir excel_report.py, investigation perf du 2026-08-04, log_test_7.txt).

Avant : _conv() rappelait _to_home_currency() -> convert_to_currency() ->
convert_to_eur()+get_rate() À CHAQUE cellule (3x par ligne), alors que
display_currency/date de conversion sont fixes pour tout l'onglet. Le taux
est maintenant calculé une seule fois avant la boucle. Ce test vérifie que
les montants convertis dans le fichier Excel produit sont numériquement
identiques à un calcul de référence utilisant l'ancien chemin
(_to_home_currency), pour un taux de change non-EUR fixe.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from tva_intracom import BuyerType, Sale, compute_all_with_vies
from tva_intracom import ecb_rates
from tva_intracom.excel_report import export_xlsx, _to_home_currency


@pytest.fixture()
def sample_results_multi():
    sales = [
        Sale("A", Decimal("100.00"), BuyerType.B2C, stock_country="FR", buyer_country="FR"),
        Sale("B", Decimal("1234.56"), BuyerType.B2C, stock_country="FR", buyer_country="DE"),
        Sale("C", Decimal("0.01"), BuyerType.B2C, stock_country="FR", buyer_country="DE"),
    ]
    return compute_all_with_vies(sales, scope_id="test-detail-conv")[0]


def test_write_details_tab_conversion_matches_reference_for_non_eur_currency(
    sample_results_multi, tmp_path
):
    output = tmp_path / "rapport.xlsx"
    fixed_rate = Decimal("4.3210")
    conv_date = date.today()

    with patch.object(ecb_rates, "get_rate", return_value=fixed_rate):
        export_xlsx(
            sample_results_multi, output, scope_id="test-detail-conv",
            display_currency="PLN",
        )

        wb = load_workbook(str(output))
        ws = wb["Detail ventes"]

        # Colonne F = "Montant HT" (voir headers de _write_details_tab).
        for row_idx, res in zip(range(2, 2 + len(sample_results_multi)), sample_results_multi):
            expected = float(_to_home_currency(res.sale.amount_ht, "PLN", conv_date))
            actual = ws.cell(row=row_idx, column=6).value
            assert actual == pytest.approx(expected), (
                f"Ligne {row_idx} : montant HT converti {actual} != référence {expected}"
            )


def test_write_details_tab_eur_passthrough_unaffected(sample_results_multi, tmp_path):
    """display_currency='EUR' (ou None) ne doit jamais appeler get_rate —
    ni dans l'ancien ni dans le nouveau chemin (montant EUR renvoyé tel quel)."""
    output = tmp_path / "rapport.xlsx"

    with patch.object(ecb_rates, "get_rate") as mock_get_rate:
        export_xlsx(
            sample_results_multi, output, scope_id="test-detail-conv-eur",
            display_currency="EUR",
        )

    mock_get_rate.assert_not_called()
