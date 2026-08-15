"""Tests de non-régression pour oss_export.py.

Ce module n'avait aucune couverture avant l'optimisation RAM (passage des
3 onglets Excel en `Workbook(write_only=True)`). Ces tests fixent le
comportement observable (feuilles, cellules fusionnées, valeurs, formules,
totaux) AVANT la réécriture, pour garantir que le passage en write_only ne
change rien au fichier produit du point de vue de l'utilisateur.
"""

from __future__ import annotations

from decimal import Decimal
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from tva_intracom import BuyerType, Sale, compute_all_with_vies
from tva_intracom.oss_export import build_oss_excel, build_b2b_excel, build_oss_csv
from tva_intracom.vies_engine import ViesResult


@pytest.fixture()
def sample_results():
    sales = [
        # Vente OSS B2C (FR -> DE), soumise TVA destination.
        Sale("A", Decimal("100"), BuyerType.B2C, stock_country="FR", buyer_country="DE"),
        Sale("B", Decimal("250"), BuyerType.B2C, stock_country="FR", buyer_country="IT"),
        # Vente B2B intracommunautaire (FR -> DE), TVA acheteur valide.
        Sale("C", Decimal("500"), BuyerType.B2B, stock_country="FR",
             buyer_country="DE", buyer_vat_valid=True, buyer_vat_number="DE123456789"),
    ]
    # `check_vat_raw` seul ne suffit pas : `validate_vat_numbers_parallel`
    # tente d'abord le cache DB (scope + global) avant tout appel HTTP, donc
    # sans SUPABASE_DB_URL (sandbox de test) toute la validation échoue et
    # la vente B2B est traitée comme B2C par sécurité (cf. tests/test_vies.py,
    # 2 tests déjà en échec pré-existant sur ce point précis — `scope_id`
    # positionnel manquant, hors-sujet ici). On mocke donc directement
    # `validate_vat_numbers_parallel`, le point d'entrée réellement appelé
    # par `compute_all_with_vies`, pour obtenir une vente B2B validée.
    with patch("tva_intracom.vies_engine.validate_vat_numbers_parallel") as mock_validate:
        mock_validate.return_value = {
            "DE123456789": ViesResult(
                valid=True, country_code="DE", vat_number="123456789", name="Firma GmbH"
            )
        }
        yield compute_all_with_vies(sales, scope_id="test-oss")[0]


def test_oss_excel_creates_expected_sheets(sample_results, tmp_path):
    output = tmp_path / "oss.xlsx"
    build_oss_excel(sample_results, output, period="2026-T2")
    wb = load_workbook(str(output))
    assert wb.sheetnames == ["OSS_Résumé", "OSS_Détail"]


def test_oss_resume_merged_title_and_headers(sample_results, tmp_path):
    output = tmp_path / "oss.xlsx"
    build_oss_excel(sample_results, output, period="2026-T2")
    wb = load_workbook(str(output))
    ws = wb["OSS_Résumé"]
    assert "A1:J1" in [str(r) for r in ws.merged_cells.ranges]
    # Ligne d'en-tête (row 3) : 10 colonnes.
    headers = [ws.cell(row=3, column=c).value for c in range(1, 11)]
    assert all(h for h in headers)


def test_oss_resume_data_rows_and_total_formula(sample_results, tmp_path):
    output = tmp_path / "oss.xlsx"
    build_oss_excel(sample_results, output, period="2026-T2")
    wb = load_workbook(str(output))
    ws = wb["OSS_Résumé"]
    # 2 pays OSS distincts (DE, IT) -> lignes 4 et 5, total ligne 6.
    countries = {ws.cell(row=r, column=1).value for r in (4, 5)}
    assert countries == {"DE", "IT"}
    total_row = 6
    assert ws.cell(row=total_row, column=4).value == "=SUM(D4:D5)"
    assert ws.cell(row=total_row, column=9).value == "=SUM(I4:I5)"


def test_oss_detail_sheet_one_row_per_sale(sample_results, tmp_path):
    output = tmp_path / "oss.xlsx"
    build_oss_excel(sample_results, output, period="2026-T2")
    wb = load_workbook(str(output))
    ws = wb["OSS_Détail"]
    # Header row 2, 2 ventes OSS (A, B) -> lignes 3 et 4, total ligne 5.
    ids = {ws.cell(row=r, column=1).value for r in (3, 4)}
    assert ids == {"A", "B"}
    assert ws.cell(row=5, column=6).value == "=SUM(F3:F4)"
    assert ws.cell(row=5, column=8).value == "=SUM(H3:H4)"


def test_b2b_excel_sheet_and_rows(sample_results, tmp_path):
    output = tmp_path / "b2b.xlsx"
    build_b2b_excel(sample_results, output, period="2026-T2")
    wb = load_workbook(str(output))
    assert wb.sheetnames == ["B2B_Recap"]
    ws = wb["B2B_Recap"]
    assert "A1:F1" in [str(r) for r in ws.merged_cells.ranges]
    # 1 vente B2B (C, sans transaction_date) -> regroupee sous le bandeau
    # "date inconnue" (ligne 4), donnee ligne 5, sous-total mois ligne 6,
    # total general ligne 7 (voir _build_b2b_recap : sous-totaux mensuels,
    # obligation DES mensuelle art. 289 B CGI).
    assert ws.cell(row=5, column=1).value == "C"
    assert ws.cell(row=5, column=3).value == "DE123456789"
    assert ws.cell(row=6, column=6).value == 500.0
    assert ws.cell(row=7, column=6).value == 500.0


def test_oss_csv_still_generated(sample_results):
    oss_bytes, b2b_bytes = build_oss_csv(sample_results, period="2026-T2")
    assert oss_bytes.startswith("\ufeff".encode("utf-8"))
    assert b2b_bytes.startswith("\ufeff".encode("utf-8"))
    assert b"DE" in oss_bytes
