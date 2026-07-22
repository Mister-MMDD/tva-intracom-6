"""Tests pour l'export Excel."""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tva_intracom import BuyerType, Sale, compute_all_with_vies
from tva_intracom.excel_report import export_xlsx
from tva_intracom.report import build_report


@pytest.fixture()
def sample_results():
    sales = [
        Sale("A", Decimal("100"), BuyerType.B2C, stock_country="FR", buyer_country="FR"),
        Sale("B", Decimal("100"), BuyerType.B2C, stock_country="FR", buyer_country="DE"),
        Sale("C", Decimal("200"), BuyerType.B2B, stock_country="FR",
             buyer_country="DE", buyer_vat_valid=True),
    ]
    return compute_all_with_vies(sales, scope_id="test")[0]


def test_export_creates_file(sample_results, tmp_path):
    output = tmp_path / "rapport.xlsx"
    result_path = export_xlsx(sample_results, output)
    assert result_path.exists()
    assert result_path.stat().st_size > 0


def test_export_has_two_sheets(sample_results, tmp_path):
    output = tmp_path / "rapport.xlsx"
    export_xlsx(sample_results, output)
    wb = load_workbook(str(output))
    assert "Recapitulatif" in wb.sheetnames
    assert "Detail ventes" in wb.sheetnames


def test_detail_sheet_rows(sample_results, tmp_path):
    output = tmp_path / "rapport.xlsx"
    export_xlsx(sample_results, output)
    wb = load_workbook(str(output))
    ws = wb["Detail ventes"]
    # Header + 3 ventes.
    assert ws.max_row == 4


def test_summary_sheet_has_total(sample_results, tmp_path):
    output = tmp_path / "rapport.xlsx"
    summary = build_report(sample_results)
    export_xlsx(sample_results, output, summary=summary)
    wb = load_workbook(str(output))
    ws = wb["Recapitulatif"]
    # Verifie qu'on retrouve le CA HT total (400.00) quelque part.
    values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert 400.0 in values


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
